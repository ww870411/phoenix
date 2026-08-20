# -*- coding: utf-8 -*-
"""
执行高温水 1、2、3、4 标段球阀基准量更新与入库：
1. 自动备份当前 high_lot_1 / high_lot_2 的高温水球阀数据；
2. 事务内删除 high_lot_1 / high_lot_2 的高温水旧球阀数据；
3. 解析《8.20 高温水1.2.3.4标球阀数量（额外）.xlsx》并将 49 条标准化球阀记录导入 tube.tube_fitting_baseline；
4. 校验与核对入库结果。
"""

import os
import sys
import json
import re
from datetime import datetime
from collections import defaultdict
import openpyxl
from sqlalchemy import text

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from backend.db.database_daily_report_25_26 import SessionLocal
from backend.projects.insulation_pipe_supply_2026.services.baseline_service import (
    ensure_baseline_tables,
    list_fitting_baselines,
)

EXCEL_PATH = os.path.join(project_root, "configs", "8.20 高温水1.2.3.4标球阀数量（额外）.xlsx")
BACKUP_DIR = os.path.join(project_root, "backend_data", "projects", "insulation_pipe_supply_2026")


def backup_existing_records(session) -> str:
    """备份当前 high_lot_1 / high_lot_2 的高温水球阀数据。"""
    rows = session.execute(text("""
        SELECT *
        FROM tube.tube_fitting_baseline
        WHERE section_1_id IN ('high_lot_1', 'high_lot_2')
          AND category = '球阀'
          AND system_type = '高温水'
        ORDER BY section_1_id, id;
    """)).mappings().all()

    backup_data = []
    for r in rows:
        item = dict(r)
        for k, v in item.items():
            if isinstance(v, datetime):
                item[k] = v.isoformat()
            elif hasattr(v, '__float__'):
                item[k] = float(v)
        backup_data.append(item)

    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup_file = os.path.join(BACKUP_DIR, "backup_ball_valves_high_lot_1_2_20260820.json")
    with open(backup_file, "w", encoding="utf-8") as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)

    print(f"📦 [1/4] 成功备份 {len(backup_data)} 条旧记录至: {backup_file}")
    return backup_file


def parse_excel_rows(excel_path: str):
    """解析 Excel 中的 49 条球阀数据并标准化为 22 维字段。"""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb["Sheet1"]

    current_section = None
    parsed_items = []

    for r in range(1, sheet.max_row + 1):
        val0 = sheet.cell(row=r, column=1).value
        val1 = sheet.cell(row=r, column=2).value
        val2 = sheet.cell(row=r, column=3).value
        val3 = sheet.cell(row=r, column=4).value
        val4 = sheet.cell(row=r, column=5).value

        if val0 and str(val0).strip().startswith("high_lot"):
            current_section = str(val0).strip()
            continue
        if val0 == "序号" or not val0:
            continue
        if isinstance(val0, (int, float)):
            form_str = str(val1).strip() if val1 else ""
            spec_str = str(val2).strip() if val2 else ""
            extra_str = str(val3).strip() if val3 else ""
            qty_val = float(val4 or 0)

            sec_id = current_section
            sys_type = "高温水"
            category = "球阀"
            if "直埋" in form_str:
                std_name = "直埋焊接球阀"
            else:
                std_name = "焊接球阀"

            model_spec = spec_str
            dn_match = re.search(r"DN\s*(\d+)", spec_str, re.I)
            main_dn = float(dn_match.group(1)) if dn_match else None

            valve_model = ""
            length_m = None
            pressure_rating = ""
            sub_model_spec = ""

            vm_match = re.search(r"([A-Za-z0-9\-]+)", extra_str)
            if vm_match:
                valve_model = vm_match.group(1)
                if "-25" in valve_model:
                    pressure_rating = "PN2.5"
                elif "-16" in valve_model:
                    pressure_rating = "PN1.6"

            len_match = re.search(r"([\d\.]+)\s*米", extra_str)
            if len_match:
                length_m = float(len_match.group(1))
                sub_model_spec = f"{length_m}米"
            else:
                sub_model_spec = ""

            unit = "套"
            raw_name = form_str
            raw_model_spec = f"{spec_str} {extra_str}".strip()
            remark = f"阀柄高: {extra_str}" if length_m else (f"型号: {extra_str}" if extra_str else "")

            parsed_items.append({
                "section_1_id": sec_id,
                "system_type": sys_type,
                "category": category,
                "standard_name": std_name,
                "model_spec": model_spec,
                "sub_model_spec": sub_model_spec,
                "unit": unit,
                "design_qty": qty_val,
                "purchase_plan_qty": qty_val,
                "main_dn": main_dn,
                "sub_dn": None,
                "angle": None,
                "bending_radius_ratio": None,
                "bending_radius_m": None,
                "valve_model": valve_model,
                "outer_diameter": None,
                "wall_thickness": None,
                "length_m": length_m,
                "pressure_rating": pressure_rating,
                "compensation_mm": None,
                "flow_direction": "",
                "raw_model_spec": raw_model_spec,
                "raw_name": raw_name,
                "remark": remark,
                "operator_name": "excel_import_20260820",
            })

    print(f"📊 [2/4] 从 Excel 成功解析出 {len(parsed_items)} 条标准化球阀数据")
    return parsed_items


def execute_import():
    """主执行逻辑：事务内先删后插。"""
    ensure_baseline_tables()
    parsed_items = parse_excel_rows(EXCEL_PATH)

    session = SessionLocal()
    try:
        # 1. 备份
        backup_existing_records(session)

        # 2. 事务删除旧记录
        print("🗑️ [3/4] 正在事务内删除 high_lot_1 与 high_lot_2 的高温水旧球阀数据...")
        del_res = session.execute(text("""
            DELETE FROM tube.tube_fitting_baseline
            WHERE section_1_id IN ('high_lot_1', 'high_lot_2')
              AND category = '球阀'
              AND system_type = '高温水';
        """))
        print(f"  -> 删除旧记录条数: {del_res.rowcount}")

        # 3. 批量插入新记录 (UPSERT)
        upsert_sql = """
            INSERT INTO tube.tube_fitting_baseline (
                section_1_id, system_type, category, standard_name, model_spec, sub_model_spec, unit,
                design_qty, purchase_plan_qty,
                main_dn, sub_dn, angle, bending_radius_ratio, bending_radius_m,
                valve_model, outer_diameter, wall_thickness, length_m,
                pressure_rating, compensation_mm, flow_direction, raw_model_spec, raw_name, remark, extra_params,
                created_by, updated_by, updated_at
            ) VALUES (
                :section_1_id, :system_type, :category, :standard_name, :model_spec, :sub_model_spec, :unit,
                :design_qty, :purchase_plan_qty,
                :main_dn, :sub_dn, :angle, :bending_radius_ratio, :bending_radius_m,
                :valve_model, :outer_diameter, :wall_thickness, :length_m,
                :pressure_rating, :compensation_mm, :flow_direction, :raw_model_spec, :raw_name, :remark, '{}'::jsonb,
                :operator_name, :operator_name, NOW()
            )
            ON CONFLICT (section_1_id, system_type, standard_name, model_spec, sub_model_spec) DO UPDATE SET
                category = EXCLUDED.category,
                unit = EXCLUDED.unit,
                design_qty = EXCLUDED.design_qty,
                purchase_plan_qty = EXCLUDED.purchase_plan_qty,
                main_dn = EXCLUDED.main_dn,
                sub_dn = EXCLUDED.sub_dn,
                angle = EXCLUDED.angle,
                bending_radius_ratio = EXCLUDED.bending_radius_ratio,
                bending_radius_m = EXCLUDED.bending_radius_m,
                valve_model = EXCLUDED.valve_model,
                outer_diameter = EXCLUDED.outer_diameter,
                wall_thickness = EXCLUDED.wall_thickness,
                length_m = EXCLUDED.length_m,
                pressure_rating = EXCLUDED.pressure_rating,
                compensation_mm = EXCLUDED.compensation_mm,
                flow_direction = EXCLUDED.flow_direction,
                raw_model_spec = EXCLUDED.raw_model_spec,
                raw_name = EXCLUDED.raw_name,
                remark = EXCLUDED.remark,
                updated_by = EXCLUDED.updated_by,
                updated_at = NOW()
        """

        for it in parsed_items:
            session.execute(text(upsert_sql), it)

        session.commit()
        print(f"✅ 成功提交事务！已写入 {len(parsed_items)} 条球阀数据。")

        # 4. 统计与核验
        print("\n🔍 [4/4] 数据库最新球阀数据核验结果:")
        summary_rows = session.execute(text("""
            SELECT section_1_id, system_type, category, COUNT(*) as cnt,
                   SUM(design_qty) as sum_design, SUM(purchase_plan_qty) as sum_plan
            FROM tube.tube_fitting_baseline
            WHERE category = '球阀'
            GROUP BY section_1_id, system_type, category
            ORDER BY system_type DESC, section_1_id ASC;
        """)).mappings().all()

        for s in summary_rows:
            print(f"  标段 {s['section_1_id']:<12} | 系统: {s['system_type']} | 类别: {s['category']} | 记录数: {s['cnt']:>2} 条 | 设计量: {s['sum_design']:>6} 套 | 计划采购量: {s['sum_plan']:>6} 套")

        total_fitting_rows = session.execute(text("SELECT COUNT(*) FROM tube.tube_fitting_baseline;")).scalar()
        print(f"\n🎉 当前 tube.tube_fitting_baseline 全表总记录数: {total_fitting_rows} 行")

    except Exception as e:
        session.rollback()
        print(f"❌ 执行发生异常，事务已安全回滚: {e}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    execute_import()
