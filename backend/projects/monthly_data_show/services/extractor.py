# -*- coding: utf-8 -*-
"""
monthly_data_show 月报入库提取服务。

第一阶段目标：
- 从上传的月报 Excel 中提取基础入库字段：
  company,item,unit,value,date,period,type,report_month
- 支持按口径（子工作表）与字段复选导出 CSV。
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from backend.projects.monthly_data_show.services.indicator_config import evaluate_formula, load_indicator_runtime_config

BLOCKED_COMPANIES = {"恒流", "天然气炉", "中水"}
ALLOWED_FIELDS = ("company", "item", "unit", "value", "date", "period", "type", "report_month")
EXTRA_EXPORT_FIELDS = ("item_transform_type", "item_transform_note")
DEFAULT_SOURCE_COLUMNS = ("本年计划", "本月计划", "本月实际", "上年同期")

ITEM_EXCLUDE_SET = {
    "本期平均供暖面积",
    "实际平均供热面积",
    "高压电量",
    "管网循环水总量",
    "管网失水量",
    "管网失水率",
    "本月内最高气温",
    "本月内最低气温",
    "本月平均气温",
    "实际检测户数（室温）",
    "检测合格户数",
    "用户室温合格率",
    "用户报修次数",
    "24小时处理次数",
    "用户报修处理及时率",
    "本月供暖小时数",
    "应按期结案总数",
    "按期结案总数",
    "信访处结率",
}

ITEM_RENAME_RULES = [
    {
        "source": "主城区一次网电厂补水量",
        "target": "一次网电厂补水量",
        "companies": ["all"],
    },
    {"source": "一次网电厂补水量", "target": "一次网电厂补水量", "companies": ["all"]},
    {"source": "高温水网电厂补水量", "target": "一次网电厂补水量", "companies": ["all"]},
    {"source": "高温水低真空电厂补水量", "target": "一次网电厂补水量", "companies": ["all"]},
    {"source": "耗外购电量", "target": "外购电量", "companies": ["all"]},
    {"source": "外购电量", "target": "外购电量", "companies": ["all"]},
    {"source": "期末供热面积", "target": "期末供暖收费面积", "companies": ["all"]},
    {"source": "单位面积耗电量", "target": "供暖电耗率", "companies": ["all"]},
    {"source": "单位面积耗水量", "target": "供暖水耗率", "companies": ["all"]},
    {"source": "单位面积耗标准煤量", "target": "供暖标准煤耗率", "companies": ["all"]},
    {"source": "锅炉设备利用率", "target": "供热设备利用率", "companies": ["all"]},
    {"source": "锅炉热效率", "target": "全厂热效率", "companies": ["all"]},
    {"source": "总热效率", "target": "全厂热效率", "companies": ["all"]},
    {"source": "炉耗油量", "target": "耗油量", "companies": ["all"]},
    {"source": "锅炉耗柴油量", "target": "耗油量", "companies": ["all"]},
    {"source": "装机总容量", "target": "锅炉设备容量", "companies": ["all"]},
    {"source": "发电煤耗率", "target": "发电标准煤耗率", "companies": ["all"]},
    {"source": "供电煤耗率", "target": "供电标准煤耗率", "companies": ["all"]},
    {"source": "供热煤耗率", "target": "供热标准煤耗率", "companies": ["all"]},
    {"source": "供热标准煤耗量", "target": "供热耗标煤量", "companies": ["all"]},
    {"source": "发电标准煤耗量", "target": "发电耗标煤量", "companies": ["all"]},
]
UNIT_NORMALIZE_RULES = [
    {"source": "米2", "target": "平方米", "exact_match": False},
    {"source": "米²", "target": "平方米", "exact_match": False},
    {"source": "/米2", "target": "/平方米", "exact_match": False},
    {"source": "/米²", "target": "/平方米", "exact_match": False},
    {"source": "千瓦时", "target": "万千瓦时", "value_divisor": 10000, "exact_match": True},
]

DEFAULT_CONSTANT_RULES = [
    {"company": "北海", "item": "发电设备容量", "unit": "万千瓦", "value": 10.0, "source_columns": ["本月实际"]},
    {"company": "香海", "item": "发电设备容量", "unit": "万千瓦", "value": 10.0, "source_columns": ["本月实际"]},
    {"company": "金州", "item": "发电设备容量", "unit": "万千瓦", "value": 3.6, "source_columns": ["本月实际"]},
    {"company": "北方", "item": "发电设备容量", "unit": "万千瓦", "value": 3.96, "source_columns": ["本月实际"]},
    {"company": "北海", "item": "锅炉设备容量", "unit": "兆瓦", "value": 577.0, "source_columns": ["本月实际"]},
    {"company": "香海", "item": "锅炉设备容量", "unit": "兆瓦", "value": 528.0, "source_columns": ["本月实际"]},
    {"company": "金州", "item": "锅炉设备容量", "unit": "兆瓦", "value": 414.0, "source_columns": ["本月实际"]},
    {"company": "北方", "item": "锅炉设备容量", "unit": "兆瓦", "value": 130.0, "source_columns": ["本月实际"]},
]

DEFAULT_SEMI_CALCULATED_RULES = [
    {
        "name": "煤折标煤量补齐",
        "companies": ["金普", "庄河"],
        "target_item": "煤折标煤量",
        "target_unit": "吨",
        "operation": "copy",
        "sources": ["耗标煤总量"],
    },
    {
        "name": "供热耗标煤量补齐",
        "companies": ["北海水炉", "金普", "庄河"],
        "target_item": "供热耗标煤量",
        "target_unit": "吨",
        "operation": "copy",
        "sources": ["耗标煤总量"],
    },
    {
        "name": "耗电量补齐",
        "companies": ["北海", "香海"],
        "target_item": "耗电量",
        "target_unit": "万千瓦时",
        "operation": "sum",
        "sources": ["综合厂用电量", "外购电量"],
        "require_any_source": True,
    },
    {
        "name": "耗电量补齐",
        "companies": ["供热公司", "金普", "庄河", "研究院", "主城区电锅炉"],
        "target_item": "耗电量",
        "target_unit": "万千瓦时",
        "operation": "copy",
        "sources": ["外购电量"],
    },
    {
        "name": "耗水量补齐",
        "companies": ["北海", "北海水炉", "香海"],
        "target_item": "耗水量",
        "target_unit": "吨",
        "operation": "copy",
        "sources": ["电厂耗水量"],
    },
    {
        "name": "热网耗水量补齐",
        "companies": ["供热公司", "金普", "庄河", "研究院", "主城区电锅炉"],
        "target_item": "热网耗水量",
        "target_unit": "吨",
        "operation": "copy",
        "sources": ["耗水量"],
    },
    {
        "name": "热网耗电量补齐",
        "companies": ["供热公司", "金普", "庄河", "研究院", "主城区电锅炉"],
        "target_item": "热网耗电量",
        "target_unit": "万千瓦时",
        "operation": "copy",
        "sources": ["外购电量"],
    },
    {
        "name": "供暖耗热量补齐",
        "companies": ["供热公司"],
        "target_item": "供暖耗热量",
        "target_unit": "吉焦",
        "operation": "copy",
        "sources": ["各热力站耗热量"],
    },
    {
        "name": "供暖耗热量补齐",
        "companies": ["金州", "北方"],
        "target_item": "供暖耗热量",
        "target_unit": "吉焦",
        "operation": "subtract",
        "sources": ["供热量", "高温水销售量"],
        "allow_missing_subtrahend_as_zero": True,
    },
    {
        "name": "供暖耗热量补齐",
        "companies": ["金普", "庄河", "研究院", "主城区电锅炉"],
        "target_item": "供暖耗热量",
        "target_unit": "吉焦",
        "operation": "copy",
        "sources": ["供热量"],
    },
]

RULES_CONFIG_CANDIDATES = [
    Path("/app/data/projects/monthly_data_show/monthly_data_show_extraction_rules.json"),
    Path("/app/backend_data/projects/monthly_data_show/monthly_data_show_extraction_rules.json"),
]
try:
    _PROJECT_ROOT = Path(__file__).resolve().parents[4]
    RULES_CONFIG_CANDIDATES.append(
        _PROJECT_ROOT / "backend_data" / "projects" / "monthly_data_show" / "monthly_data_show_extraction_rules.json"
    )
except Exception:
    pass

_BASE_RULES_SNAPSHOT = {
    "blocked_companies": sorted(BLOCKED_COMPANIES),
    "default_source_columns": list(DEFAULT_SOURCE_COLUMNS),
    "item_exclude_set": sorted(ITEM_EXCLUDE_SET),
    "item_rename_rules": [dict(rule) for rule in ITEM_RENAME_RULES],
    "unit_normalize_rules": [dict(rule) for rule in UNIT_NORMALIZE_RULES],
    "default_constant_rules": [dict(x) for x in DEFAULT_CONSTANT_RULES],
    "semi_calculated_rules": [dict(x) for x in DEFAULT_SEMI_CALCULATED_RULES],
}
SEMI_CALCULATED_RULES = [dict(x) for x in DEFAULT_SEMI_CALCULATED_RULES]
UNIT_NORMALIZE_RULES_RUNTIME = [dict(x) for x in UNIT_NORMALIZE_RULES]


def _load_extraction_rules_from_config() -> Dict[str, Any]:
    for path in RULES_CONFIG_CANDIDATES:
        try:
            if not path.exists():
                continue
            payload = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                return payload
        except Exception:
            continue
    return {}


def _refresh_extraction_rules() -> None:
    global BLOCKED_COMPANIES
    global DEFAULT_SOURCE_COLUMNS
    global ITEM_EXCLUDE_SET
    global ITEM_RENAME_RULES
    global UNIT_NORMALIZE_RULES_RUNTIME
    global DEFAULT_CONSTANT_RULES
    global SEMI_CALCULATED_RULES

    cfg = _load_extraction_rules_from_config()
    BLOCKED_COMPANIES = set(cfg.get("blocked_companies") or _BASE_RULES_SNAPSHOT["blocked_companies"])
    DEFAULT_SOURCE_COLUMNS = tuple(cfg.get("default_source_columns") or _BASE_RULES_SNAPSHOT["default_source_columns"])
    ITEM_EXCLUDE_SET = set(cfg.get("item_exclude_set") or _BASE_RULES_SNAPSHOT["item_exclude_set"])
    raw_rename_rules = cfg.get("item_rename_rules")
    if isinstance(raw_rename_rules, list):
        normalized_rename_rules: List[Dict[str, Any]] = []
        for idx, raw_rule in enumerate(raw_rename_rules):
            if not isinstance(raw_rule, dict):
                continue
            companies = [str(x).strip() for x in (raw_rule.get("companies") or []) if str(x).strip()]
            if not companies:
                companies = ["all"]
            source = str(raw_rule.get("source") or "").strip()
            target = str(raw_rule.get("target") or "").strip()
            if source and target:
                normalized_rename_rules.append(
                    {
                        "id": str(raw_rule.get("id") or "").strip() or f"item_rename::{idx + 1}",
                        "source": source,
                        "target": target,
                        "companies": companies,
                    }
                )
                continue
        ITEM_RENAME_RULES = normalized_rename_rules or [dict(rule) for rule in _BASE_RULES_SNAPSHOT["item_rename_rules"]]
    else:
        ITEM_RENAME_RULES = [dict(rule) for rule in _BASE_RULES_SNAPSHOT["item_rename_rules"]]
    for idx, rule in enumerate(ITEM_RENAME_RULES):
        if not isinstance(rule, dict):
            continue
        rule["id"] = str(rule.get("id") or "").strip() or f"item_rename::{idx + 1}"
    raw_unit_rules = cfg.get("unit_normalize_rules")
    if isinstance(raw_unit_rules, list):
        normalized_unit_rules: List[Dict[str, Any]] = []
        for raw_rule in raw_unit_rules:
            if not isinstance(raw_rule, dict):
                continue
            source = str(raw_rule.get("source") or "").strip()
            target = str(raw_rule.get("target") or "").strip()
            if not source or not target:
                continue
            rule: Dict[str, Any] = {
                "source": source,
                "target": target,
                "exact_match": bool(raw_rule.get("exact_match")),
            }
            divisor = raw_rule.get("value_divisor")
            if divisor is not None:
                try:
                    rule["value_divisor"] = float(divisor)
                except (TypeError, ValueError):
                    pass
            normalized_unit_rules.append(rule)
        UNIT_NORMALIZE_RULES_RUNTIME = normalized_unit_rules or [dict(rule) for rule in _BASE_RULES_SNAPSHOT["unit_normalize_rules"]]
    else:
        UNIT_NORMALIZE_RULES_RUNTIME = [dict(rule) for rule in _BASE_RULES_SNAPSHOT["unit_normalize_rules"]]
    for idx, rule in enumerate(UNIT_NORMALIZE_RULES_RUNTIME):
        if not isinstance(rule, dict):
            continue
        rule["id"] = str(rule.get("id") or "").strip() or f"unit_normalize::{idx + 1}"
    raw_constants = cfg.get("default_constant_rules") or _BASE_RULES_SNAPSHOT["default_constant_rules"]
    DEFAULT_CONSTANT_RULES = [dict(x) for x in raw_constants if isinstance(x, dict)]
    raw_semi = cfg.get("semi_calculated_rules") or _BASE_RULES_SNAPSHOT["semi_calculated_rules"]
    normalized_semi: List[Dict[str, Any]] = []
    for idx, raw in enumerate(raw_semi):
        if not isinstance(raw, dict):
            continue
        row = dict(raw)
        rule_id = str(row.get("id") or "").strip() or f"semi_rule_{idx + 1}"
        row["id"] = rule_id
        row["name"] = str(row.get("name") or "").strip() or rule_id
        normalized_semi.append(row)
    SEMI_CALCULATED_RULES = normalized_semi


def get_extraction_rule_options() -> List[Dict[str, Any]]:
    _refresh_extraction_rules()

    exclude_children = [
        {
            "id": f"item_exclude::{item}",
            "name": str(item),
            "description": f"剔除指标：{item}",
            "enabled_default": True,
            "item": str(item),
        }
        for item in sorted(ITEM_EXCLUDE_SET)
    ]

    rename_children: List[Dict[str, Any]] = []
    for idx, rule in enumerate(ITEM_RENAME_RULES):
        companies = [str(x).strip() for x in (rule.get("companies") or []) if str(x).strip()]
        source = str(rule.get("source") or "").strip()
        target = str(rule.get("target") or "").strip()
        if not source or not target:
            continue
        company_text = "/".join(companies) if companies else "全部"
        rename_children.append(
            {
                "id": str(rule.get("id") or "").strip() or f"item_rename::{idx + 1}",
                "name": f"{source} → {target}",
                "description": f"口径={company_text}，将 {source} 统一为 {target}",
                "enabled_default": True,
                "source": source,
                "target": target,
                "companies": companies,
            }
        )

    unit_children: List[Dict[str, Any]] = []
    for idx, rule in enumerate(UNIT_NORMALIZE_RULES_RUNTIME):
        source = str(rule.get("source") or "").strip()
        target = str(rule.get("target") or "").strip()
        if not source or not target:
            continue
        exact_match = bool(rule.get("exact_match"))
        divisor = rule.get("value_divisor")
        desc_parts = [f"单位 {source} → {target}"]
        desc_parts.append(f"完全匹配：{'是' if exact_match else '否'}")
        if divisor not in (None, ""):
            desc_parts.append(f"数值除以 {divisor}")
        unit_children.append(
            {
                "id": str(rule.get("id") or "").strip() or f"unit_normalize::{idx + 1}",
                "name": f"{source} → {target}",
                "description": "，".join(desc_parts),
                "enabled_default": True,
                "source": source,
                "target": target,
                "exact_match": exact_match,
                "value_divisor": divisor,
            }
        )

    def _describe_semi_rule(rule: Dict[str, Any]) -> str:
        companies = [str(x).strip() for x in (rule.get("companies") or []) if str(x).strip()]
        target_item = str(rule.get("target_item") or "").strip()
        target_unit = str(rule.get("target_unit") or "").strip()
        operation = str(rule.get("operation") or "").strip().lower()
        sources = [str(x).strip() for x in (rule.get("sources") or []) if str(x).strip()]
        formula = str(rule.get("formula") or "").strip()
        company_text = f"口径={ '/'.join(companies) if companies else '全部' }"
        if formula:
            expr = formula[1:].strip() if formula.startswith("=") else formula
        elif operation == "copy" and sources:
            expr = sources[0]
        elif operation == "sum" and sources:
            expr = " + ".join(sources)
        elif operation == "subtract" and sources:
            expr = " - ".join(sources)
        else:
            expr = "按配置规则计算"
        if target_item:
            return f"{company_text}，{target_item} = {expr}，单位={target_unit or '按配置'}"
        return f"{company_text}，规则表达式={expr}"

    semi_children = [
        {
            "id": str(rule.get("id") or ""),
            "name": str(rule.get("name") or ""),
            "description": str(rule.get("description") or _describe_semi_rule(rule)),
            "enabled_default": True,
        }
        for rule in SEMI_CALCULATED_RULES
        if str(rule.get("id") or "").strip()
    ]

    return [
        {
            "id": "item_exclude",
            "name": "指标剔除",
            "description": "可展开勾选 item_exclude_set 中的具体剔除项",
            "enabled_default": True,
            "children": exclude_children,
        },
        {
            "id": "item_rename",
            "name": "指标重命名",
            "description": "可展开勾选 item_rename_rules 中的具体重命名规则",
            "enabled_default": True,
            "children": rename_children,
        },
        {
            "id": "unit_normalize",
            "name": "计量单位转换",
            "description": "可展开勾选 unit_normalize_rules 中的具体单位转换规则",
            "enabled_default": True,
            "children": unit_children,
        },
        {
            "id": "semi_calculated",
            "name": "半计算规则",
            "description": "可展开勾选 semi_calculated_rules 中的具体补齐规则",
            "enabled_default": True,
            "children": semi_children,
        },
    ]


def _clean_text(value: object) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", "", text)


def _resolve_item_rename_map(company: str, rename_rules: Optional[Sequence[Dict[str, Any]]] = None) -> Dict[str, str]:
    resolved: Dict[str, str] = {}
    company_text = str(company or "").strip()
    active_rules = list(rename_rules or ITEM_RENAME_RULES)
    for rule in active_rules:
        companies = {str(x).strip() for x in (rule.get("companies") or []) if str(x).strip()}
        if "all" not in companies and company_text not in companies:
            continue
        source = str(rule.get("source") or "").strip()
        target = str(rule.get("target") or "").strip()
        if source and target:
            resolved[source] = target
    return resolved


def _append_transform_meta(
    row: Dict[str, object],
    transform_type: str = "",
    transform_note: str = "",
) -> None:
    type_parts = [str(x).strip() for x in str(row.get("item_transform_type") or "").split("；") if str(x).strip()]
    note_parts = [str(x).strip() for x in str(row.get("item_transform_note") or "").split("；") if str(x).strip()]
    if transform_type and transform_type not in type_parts:
        type_parts.append(transform_type)
    if transform_note and transform_note not in note_parts:
        note_parts.append(transform_note)
    row["item_transform_type"] = "；".join(type_parts)
    row["item_transform_note"] = "；".join(note_parts)


def _normalize_item(
    value: object,
    use_rename: bool = True,
    company: str = "",
    rename_rules: Optional[Sequence[Dict[str, Any]]] = None,
) -> Tuple[str, str, str]:
    raw = _clean_text(value)
    if not raw:
        return "", "", ""
    normalized = raw.replace("其中：", "")
    transform_type = ""
    transform_note = ""
    if use_rename:
        rename_map = _resolve_item_rename_map(company, rename_rules=rename_rules)
        mapped = rename_map.get(normalized, normalized)
        if mapped != normalized:
            transform_type = "指标更名"
            transform_note = f"{normalized}→{mapped}"
            normalized = mapped
        else:
            token = (
                normalized.replace("、", "")
                .replace("，", "")
                .replace(",", "")
                .replace("/", "")
            )
            mapped_token = rename_map.get(token, normalized)
            if mapped_token != normalized:
                transform_type = "指标更名"
                transform_note = f"{normalized}→{mapped_token}"
                normalized = mapped_token
    return normalized, transform_type, transform_note


def _normalize_unit(value: object, unit_rules: Optional[Sequence[Dict[str, Any]]] = None) -> Tuple[str, str, str]:
    raw = _clean_text(value)
    if not raw:
        return "", "", ""
    normalized = raw
    note_parts: List[str] = []
    active_rules = list(UNIT_NORMALIZE_RULES_RUNTIME if unit_rules is None else unit_rules)
    for rule in active_rules:
        source = str(rule.get("source") or "").strip()
        target = str(rule.get("target") or "").strip()
        exact_match = bool(rule.get("exact_match"))
        if not source or not target:
            continue
        if exact_match:
            if normalized != source:
                continue
            updated = target
        else:
            if source not in normalized:
                continue
            updated = normalized.replace(source, target)
        if updated != normalized:
            note_parts.append(f"{source}→{target}")
            normalized = updated
    transform_type = "单位转换" if note_parts else ""
    return normalized, transform_type, "；".join(note_parts)


SEMI_FORMULA_QUOTED_TOKEN_PATTERN = re.compile(r'"([^"]+)"|\'([^\']+)\'')
SEMI_FORMULA_TOKEN_PATTERN = re.compile(r"\{\{\s*([^{}]+?)\s*\}\}")


def _coerce_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_semicalculated_formula(formula: object) -> str:
    text = str(formula or "").strip()
    if text.startswith("="):
        text = text[1:].strip()
    if not text:
        return ""

    def _replace(match: re.Match) -> str:
        token = str(match.group(1) or match.group(2) or "").strip()
        return f"{{{{{token}}}}}" if token else ""

    return SEMI_FORMULA_QUOTED_TOKEN_PATTERN.sub(_replace, text)


def _extract_semicalculated_formula_tokens(formula: str) -> List[str]:
    seen = set()
    tokens: List[str] = []
    for raw in SEMI_FORMULA_TOKEN_PATTERN.findall(formula or ""):
        token = str(raw or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)
        tokens.append(token)
    return tokens


def _resolve_formula_token_scope(current_company: str, raw_token: str) -> Tuple[str, str]:
    token = str(raw_token or "").strip()
    if not token:
        return current_company, ""
    for sep in ("::", "：", ":"):
        if sep not in token:
            continue
        company_text, item_text = token.split(sep, 1)
        company_name = str(company_text or "").strip() or current_company
        item_name = str(item_text or "").strip()
        if company_name in {"本口径", "当前口径", "self", "current"}:
            company_name = current_company
        return company_name, item_name
    return current_company, token


def _evaluate_semicalculated_formula(
    formula: object,
    company: str,
    win: Tuple[str, str, str, str],
    value_getter,
    allow_missing_as_zero: bool = False,
) -> Tuple[Optional[float], str]:
    normalized_formula = _normalize_semicalculated_formula(formula)
    if not normalized_formula:
        return None, ""

    context: Dict[str, float] = {}
    missing_tokens: List[str] = []
    for token in _extract_semicalculated_formula_tokens(normalized_formula):
        source_company, item_name = _resolve_formula_token_scope(company, token)
        token_value = value_getter(source_company, win, item_name)
        if token_value is None:
            if allow_missing_as_zero:
                context[token] = 0.0
            else:
                missing_tokens.append(token)
            continue
        context[token] = float(token_value)
    if missing_tokens:
        return None, normalized_formula
    try:
        return float(evaluate_formula(normalized_formula, context)), normalized_formula
    except Exception:
        return None, normalized_formula


def _normalize_value(raw_unit: str, unit: str, value: object) -> object:
    number = _coerce_number(value)
    if number is None:
        return str(value).strip()
    for rule in UNIT_NORMALIZE_RULES_RUNTIME:
        source = str(rule.get("source") or "").strip()
        target = str(rule.get("target") or "").strip()
        divisor = rule.get("value_divisor")
        exact_match = bool(rule.get("exact_match"))
        if not source or not target:
            continue
        if exact_match:
            normalized_by_rule = target if raw_unit == source else raw_unit
        else:
            normalized_by_rule = raw_unit.replace(source, target) if source in raw_unit else raw_unit
        if normalized_by_rule == unit and divisor not in (None, 0, 0.0):
            return round(number / float(divisor), 8)
    return round(number, 8)


def _build_period_meta(report_year: int, report_month: int, src_col: str) -> Dict[str, str]:
    if src_col == "本年计划":
        return {
            "date": f"{report_year}-01-01",
            "period": "year",
            "type": "plan",
        }
    if src_col == "本月计划":
        return {
            "date": f"{report_year}-{report_month:02d}-01",
            "period": "month",
            "type": "plan",
        }
    if src_col == "本月实际":
        return {
            "date": f"{report_year}-{report_month:02d}-01",
            "period": "month",
            "type": "real",
        }
    return {
        "date": f"{report_year - 1}-{report_month:02d}-01",
        "period": "month",
        "type": "real",
    }


def _build_report_month_text(report_year: int, report_month: int) -> str:
    return f"{report_year}-{report_month:02d}-01"


def _apply_semicalculated_completion_rules(
    rows: List[Dict[str, object]],
    enabled_rule_ids: Optional[set[str]] = None,
) -> Dict[str, int]:
    """
    按“2.28 月报数据库化配置文件”第四部分补齐半计算指标。
    规则原则：
    - 仅覆盖规则中明确指定的口径；
    - 未指定口径沿用原值；
    - 同一口径、同一时期下若命中规则，则以规则结果重写目标指标。
    """

    def _window_key(row: Dict[str, object]) -> Tuple[str, str, str, str]:
        return (
            str(row.get("date") or "").strip(),
            str(row.get("period") or "").strip(),
            str(row.get("type") or "").strip(),
            str(row.get("report_month") or "").strip(),
        )

    def _row_key(row: Dict[str, object]) -> Tuple[str, str, str, str, str, str]:
        return (
            str(row.get("company") or "").strip(),
            str(row.get("item") or "").strip(),
            str(row.get("date") or "").strip(),
            str(row.get("period") or "").strip(),
            str(row.get("type") or "").strip(),
            str(row.get("report_month") or "").strip(),
        )

    company_window_values: Dict[Tuple[str, Tuple[str, str, str, str]], Dict[str, float]] = {}
    row_index_by_key: Dict[Tuple[str, str, str, str, str, str], int] = {}
    for idx, row in enumerate(rows):
        company = str(row.get("company") or "").strip()
        item = str(row.get("item") or "").strip()
        win = _window_key(row)
        row_index_by_key[_row_key(row)] = idx
        numeric = _coerce_number(row.get("value"))
        if not company or not item or numeric is None:
            continue
        slot = company_window_values.setdefault((company, win), {})
        slot[item] = float(numeric)

    def _value_of(company: str, win: Tuple[str, str, str, str], item: str) -> Optional[float]:
        return company_window_values.get((company, win), {}).get(item)

    def _upsert_value(
        company: str,
        win: Tuple[str, str, str, str],
        item: str,
        unit: str,
        value: float,
        transform_type: str = "",
        transform_note: str = "",
    ) -> None:
        new_row = {
            "company": company,
            "item": item,
            "unit": unit,
            "value": round(float(value), 8),
            "date": win[0],
            "period": win[1],
            "type": win[2],
            "report_month": win[3],
            "item_transform_type": "",
            "item_transform_note": "",
        }
        _append_transform_meta(new_row, transform_type=transform_type, transform_note=transform_note)
        key = (
            company,
            item,
            win[0],
            win[1],
            win[2],
            win[3],
        )
        old_idx = row_index_by_key.get(key)
        if old_idx is None:
            row_index_by_key[key] = len(rows)
            rows.append(new_row)
        else:
            rows[old_idx] = new_row
        company_window_values.setdefault((company, win), {})[item] = float(value)

    details: Dict[str, int] = {}
    rule_name_map: Dict[str, str] = {}
    for rule in SEMI_CALCULATED_RULES:
        rule_id = str(rule.get("id") or "").strip()
        rule_name = str(rule.get("name") or "").strip() or rule_id or "未命名规则"
        if not rule_id:
            continue
        details.setdefault(rule_id, 0)
        rule_name_map[rule_id] = rule_name
    all_company_windows = list(company_window_values.keys())

    for company, win in all_company_windows:
        for rule in SEMI_CALCULATED_RULES:
            rule_id = str(rule.get("id") or "").strip()
            if not rule_id or (enabled_rule_ids is not None and rule_id not in enabled_rule_ids):
                continue
            companies = {str(x).strip() for x in (rule.get("companies") or []) if str(x).strip()}
            if companies and company not in companies:
                continue
            target_item = str(rule.get("target_item") or "").strip()
            target_unit = str(rule.get("target_unit") or "").strip()
            operation = str(rule.get("operation") or "").strip().lower()
            sources = [str(x).strip() for x in (rule.get("sources") or []) if str(x).strip()]
            formula = str(rule.get("formula") or "").strip()
            if not target_item or not target_unit:
                continue

            value: Optional[float] = None
            expr = "规则计算"
            if formula:
                value, normalized_formula = _evaluate_semicalculated_formula(
                    formula=formula,
                    company=company,
                    win=win,
                    value_getter=_value_of,
                    allow_missing_as_zero=bool(rule.get("allow_missing_formula_sources_as_zero", False)),
                )
                expr = formula[1:].strip() if formula.startswith("=") else formula
                if normalized_formula:
                    expr = normalized_formula
            else:
                if not operation or not sources:
                    continue
                if operation == "copy":
                    value = _value_of(company, win, sources[0])
                elif operation == "sum":
                    require_any = bool(rule.get("require_any_source", False))
                    parts = [_value_of(company, win, src) for src in sources]
                    if require_any and not any(v is not None for v in parts):
                        continue
                    if (not require_any) and any(v is None for v in parts):
                        continue
                    value = float(sum(float(v or 0.0) for v in parts))
                elif operation == "subtract":
                    minuend = _value_of(company, win, sources[0])
                    if minuend is None:
                        continue
                    remainder = float(minuend)
                    allow_missing_sub = bool(rule.get("allow_missing_subtrahend_as_zero", True))
                    for src in sources[1:]:
                        sub = _value_of(company, win, src)
                        if sub is None and not allow_missing_sub:
                            remainder = None
                            break
                        remainder -= float(sub or 0.0)
                    value = remainder
                if operation == "copy":
                    expr = sources[0]
                elif operation == "sum":
                    expr = " + ".join(sources)
                elif operation == "subtract":
                    expr = " - ".join(sources)
            if value is None:
                continue
            note = f"{expr}→{target_item}"
            _upsert_value(
                company,
                win,
                target_item,
                target_unit,
                float(value),
                transform_type="半计算",
                transform_note=note,
            )
            details[rule_id] = int(details.get(rule_id, 0)) + 1

    named_details: Dict[str, int] = {}
    for rule_id, count in details.items():
        rule_name = rule_name_map.get(rule_id, rule_id)
        named_details[rule_name] = int(named_details.get(rule_name, 0)) + int(count)
    return named_details



def parse_report_period_from_filename(filename: str) -> Tuple[int, int]:
    text = str(filename or "")
    matched = re.search(r"(\d{2})\.(\d{1,2})", text)
    if matched:
        year = 2000 + int(matched.group(1))
        month = int(matched.group(2))
        month = max(1, min(12, month))
        return year, month
    now = datetime.now()
    return now.year, now.month


def get_default_constant_rules() -> List[Dict[str, object]]:
    _refresh_extraction_rules()
    return [dict(item) for item in DEFAULT_CONSTANT_RULES]


def normalize_constant_rules(raw_rules: Optional[Sequence[Dict[str, object]]]) -> List[Dict[str, object]]:
    _refresh_extraction_rules()
    rules: List[Dict[str, object]] = []
    for raw in raw_rules or []:
        if not isinstance(raw, dict):
            continue
        company = str(raw.get("company") or "").strip()
        item, item_transform_type, item_transform_note = _normalize_item(raw.get("item"), use_rename=True, company=company)
        unit, unit_transform_type, unit_transform_note = _normalize_unit(raw.get("unit"))
        if not company or not item:
            continue
        value = _coerce_number(raw.get("value"))
        if value is None:
            continue
        src_cols_raw = raw.get("source_columns")
        source_columns: List[str] = []
        if isinstance(src_cols_raw, list):
            for col in src_cols_raw:
                col_name = str(col or "").strip()
                if col_name in DEFAULT_SOURCE_COLUMNS and col_name not in source_columns:
                    source_columns.append(col_name)
        if not source_columns:
            fallback_col = str(raw.get("source_column") or "本月实际").strip()
            if fallback_col in DEFAULT_SOURCE_COLUMNS:
                source_columns = [fallback_col]
            else:
                source_columns = ["本月实际"]
        rules.append(
            {
                "company": company,
                "item": item,
                "unit": unit,
                "value": round(float(value), 8),
                "source_columns": source_columns,
                "item_transform_type": "；".join(
                    [x for x in [item_transform_type, unit_transform_type] if x]
                ),
                "item_transform_note": "；".join(
                    [x for x in [item_transform_note, unit_transform_note] if x]
                ),
            }
        )
    return rules


def _find_header_row_and_columns(sheet) -> Tuple[int, Dict[str, int]]:
    required = {"项目", "计量单位", "本年计划", "本月计划", "上年同期", "本月实际"}
    for row_index in range(1, min(8, sheet.max_row + 1)):
        mapping: Dict[str, int] = {}
        for col_index in range(1, sheet.max_column + 1):
            label = _clean_text(sheet.cell(row=row_index, column=col_index).value)
            if not label:
                continue
            for target in required:
                if label == target:
                    mapping[target] = col_index
        if required.issubset(set(mapping.keys())):
            return row_index, mapping
    raise ValueError(f"子表 {sheet.title} 未找到标准表头（项目/计量单位/计划/实际）")


def _iter_valid_sheets(workbook, selected_companies: Optional[Sequence[str]] = None):
    selected = {str(x).strip() for x in (selected_companies or []) if str(x).strip()}
    for name in workbook.sheetnames:
        company = str(name).strip()
        if not company or company in BLOCKED_COMPANIES:
            continue
        if selected and company not in selected:
            continue
        yield workbook[company]


def get_company_options(file_bytes: bytes) -> List[str]:
    _refresh_extraction_rules()
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    companies = [name for name in workbook.sheetnames if str(name).strip() and str(name).strip() not in BLOCKED_COMPANIES]
    workbook.close()
    return companies


def extract_rows(
    file_bytes: bytes,
    filename: str,
    selected_companies: Optional[Sequence[str]] = None,
    selected_source_columns: Optional[Sequence[str]] = None,
    selected_rule_ids: Optional[Sequence[str]] = None,
    constants_enabled: bool = False,
    constant_rules: Optional[Sequence[Dict[str, object]]] = None,
    report_year: Optional[int] = None,
    report_month: Optional[int] = None,
) -> Tuple[List[Dict[str, object]], Dict[str, object]]:
    _refresh_extraction_rules()
    runtime_cfg = load_indicator_runtime_config()
    calculated_item_set = set(runtime_cfg.get("calculated_item_set") or set())
    parsed_year, parsed_month = parse_report_period_from_filename(filename)
    if report_year is None:
        report_year = parsed_year
    if report_month is None:
        report_month = parsed_month
    report_month = max(1, min(12, int(report_month)))
    report_year = int(report_year)
    report_month_text = _build_report_month_text(report_year, report_month)
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    rows: List[Dict[str, object]] = []
    per_company_count: Dict[str, int] = {}

    available_rule_groups = get_extraction_rule_options()
    child_rule_ids: set[str] = set()
    parent_to_children: Dict[str, set[str]] = {}
    for group in available_rule_groups:
        parent_id = str(group.get("id") or "").strip()
        children = [x for x in (group.get("children") or []) if isinstance(x, dict)]
        ids = {str(child.get("id") or "").strip() for child in children if str(child.get("id") or "").strip()}
        if parent_id:
            parent_to_children[parent_id] = ids
        child_rule_ids.update(ids)

    raw_selected_rules = {str(x).strip() for x in (selected_rule_ids or []) if str(x).strip()}
    if not raw_selected_rules:
        selected_rules = set(child_rule_ids)
    else:
        selected_rules = set()
        for rule_id in raw_selected_rules:
            if rule_id in parent_to_children:
                selected_rules.update(parent_to_children.get(rule_id) or set())
            else:
                selected_rules.add(rule_id)
    selected_rules = {rule_id for rule_id in selected_rules if rule_id in child_rule_ids}

    selected_exclude_items = {
        str(rule_id).split("::", 1)[1]
        for rule_id in selected_rules
        if str(rule_id).startswith("item_exclude::") and "::" in str(rule_id)
    }
    selected_rename_rule_ids = {
        str(rule_id)
        for rule_id in selected_rules
        if str(rule_id).startswith("item_rename::")
    }
    selected_unit_rule_ids = {
        str(rule_id)
        for rule_id in selected_rules
        if str(rule_id).startswith("unit_normalize::")
    }
    active_rename_rules = [
        dict(rule)
        for rule in ITEM_RENAME_RULES
        if str(rule.get("id") or "").strip() in selected_rename_rule_ids
    ]
    active_unit_rules = [
        dict(rule)
        for rule in UNIT_NORMALIZE_RULES_RUNTIME
        if str(rule.get("id") or "").strip() in selected_unit_rule_ids
    ]
    use_item_exclude = bool(selected_exclude_items)
    use_item_rename = bool(active_rename_rules)
    rename_hit_count = 0
    exclude_hit_count = 0

    source_columns = tuple(DEFAULT_SOURCE_COLUMNS)
    selected_col_set = {str(x).strip() for x in (selected_source_columns or []) if str(x).strip()}
    if selected_col_set:
        source_columns = tuple(col for col in source_columns if col in selected_col_set)
    try:
        for sheet in _iter_valid_sheets(workbook, selected_companies=selected_companies):
            company = str(sheet.title).strip()
            header_row, col_map = _find_header_row_and_columns(sheet)
            count_before = len(rows)
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                item_raw = sheet.cell(row=row_idx, column=col_map["项目"]).value
                item_for_exclude, _, _ = _normalize_item(item_raw, use_rename=False, company=company)
                if not item_for_exclude:
                    continue
                if use_item_exclude and item_for_exclude in selected_exclude_items:
                    exclude_hit_count += 1
                    continue
                unit_raw = sheet.cell(row=row_idx, column=col_map["计量单位"]).value
                item, item_transform_type, item_transform_note = _normalize_item(
                    item_raw,
                    use_rename=use_item_rename,
                    company=company,
                    rename_rules=active_rename_rules,
                )
                raw_unit = _clean_text(unit_raw)
                unit, unit_transform_type, unit_transform_note = _normalize_unit(unit_raw, unit_rules=active_unit_rules)
                if not item:
                    continue
                if item_transform_type == "指标更名":
                    rename_hit_count += 1
                if item in calculated_item_set:
                    continue
                for src_col in source_columns:
                    value_cell = sheet.cell(row=row_idx, column=col_map[src_col]).value
                    value = _normalize_value(raw_unit, unit, value_cell)
                    meta = _build_period_meta(report_year, report_month, src_col)
                    row = {
                        "company": company,
                        "item": item,
                        "unit": unit,
                        "value": value,
                        "date": meta["date"],
                        "period": meta["period"],
                        "type": meta["type"],
                        "report_month": report_month_text,
                        "item_transform_type": "",
                        "item_transform_note": "",
                    }
                    _append_transform_meta(row, transform_type=item_transform_type, transform_note=item_transform_note)
                    _append_transform_meta(row, transform_type=unit_transform_type, transform_note=unit_transform_note)
                    rows.append(row)
            per_company_count[company] = len(rows) - count_before
    finally:
        workbook.close()

    stats = {
        "report_year": report_year,
        "report_month": report_month,
        "total_rows": len(rows),
        "company_rows": per_company_count,
        "item_exclude_hits": exclude_hit_count,
        "item_rename_hits": rename_hit_count,
        "selected_rule_ids": sorted(selected_rules),
    }

    constants_injected = 0
    if constants_enabled:
        normalized_rules = normalize_constant_rules(constant_rules)
        selected_company_set = {str(x).strip() for x in (selected_companies or []) if str(x).strip()}
        selected_source_set = set(source_columns)
        row_index_by_key: Dict[Tuple[str, str, str, str, str], int] = {}
        for idx, row in enumerate(rows):
            key = (
                str(row.get("company") or "").strip(),
                str(row.get("item") or "").strip(),
                str(row.get("date") or "").strip(),
                str(row.get("period") or "").strip(),
                str(row.get("type") or "").strip(),
            )
            row_index_by_key[key] = idx

        for rule in normalized_rules:
            company = str(rule["company"])
            if company in BLOCKED_COMPANIES:
                continue
            if selected_company_set and company not in selected_company_set:
                continue
            for source_column in rule.get("source_columns", []):
                source_column = str(source_column)
                if selected_source_set and source_column not in selected_source_set:
                    continue
                meta = _build_period_meta(report_year, report_month, source_column)
                new_row = {
                    "company": company,
                    "item": str(rule["item"]),
                    "unit": str(rule["unit"]),
                    "value": round(float(rule["value"]), 8),
                    "date": meta["date"],
                    "period": meta["period"],
                    "type": meta["type"],
                    "report_month": report_month_text,
                    "item_transform_type": "",
                    "item_transform_note": "",
                }
                _append_transform_meta(new_row, transform_type=str(rule.get("item_transform_type") or ""), transform_note=str(rule.get("item_transform_note") or ""))
                _append_transform_meta(new_row, transform_type="常量注入", transform_note="固定值注入")
                key = (new_row["company"], new_row["item"], new_row["date"], new_row["period"], new_row["type"])
                old_idx = row_index_by_key.get(key)
                if old_idx is None:
                    row_index_by_key[key] = len(rows)
                    rows.append(new_row)
                    constants_injected += 1
                else:
                    rows[old_idx] = new_row
                    constants_injected += 1

    stats["constants_injected"] = constants_injected
    semi_details = _apply_semicalculated_completion_rules(rows, enabled_rule_ids=selected_rules)
    stats["semi_calculated_details"] = semi_details
    stats["semi_calculated_completed"] = int(sum(semi_details.values()))
    stats["total_rows"] = len(rows)
    return rows, stats


def filter_fields(rows: Iterable[Dict[str, object]], selected_fields: Sequence[str]) -> List[Dict[str, object]]:
    field_set = [field for field in ALLOWED_FIELDS if field in set(selected_fields)]
    if not field_set:
        field_set = list(ALLOWED_FIELDS)
    field_set.extend([field for field in EXTRA_EXPORT_FIELDS if field not in field_set])
    return [{field: row.get(field, "") for field in field_set} for row in rows]
