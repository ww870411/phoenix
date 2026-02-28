# -*- coding: utf-8 -*-
"""
monthly_data_pull 导表引擎（openpyxl 版）。

说明：
- 目标是替代原 xlwings 链路，先实现可在线运行的基础能力；
- 当前版本支持：
  1) 映射表解析（源/目标键、分组、源 sheet 需求）
  2) 本月值读取（单元格或简单表达式）
  3) 累计粘贴动作（推荐动作包含“粘贴”且“累计”）
  4) 输出目标文件副本
"""

from __future__ import annotations

import datetime
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook


COL_SRC_KEY = "子公司月报表源文件"
COL_SRC_SHEET = "源sheet"
COL_SRC_CELL = "源本月"
COL_SRC_ACC = "源累计"
COL_TGT_KEY = "目标文件"
COL_TGT_SHEET = "目标sheet"
COL_TGT_CELL = "目标本月"
COL_TGT_ACC = "目标累计"
COL_ACTION = "推荐动作"
COL_INDICATOR_NAME = "子公司月报表指标名称"

REQUIRED_COLUMNS = [COL_SRC_KEY, COL_TGT_KEY]
CELL_REF_PATTERN = re.compile(r"[A-Za-z]{1,3}\d{1,7}")
SIMPLE_CELL_PATTERN = re.compile(r"^[A-Za-z]{1,3}\d{1,7}$")
SHEET_CELL_REF_PATTERN = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_\-\u4e00-\u9fff ]+))!\$?([A-Za-z]{1,3})\$?(\d{1,7})"
)


@dataclass
class StoredFileInfo:
    bucket: str
    stored_name: str
    sheet: str


def _safe_strip(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_indicator_name(row: Dict[str, Any]) -> str:
    # 按用户约定，优先且默认只取“子公司月报表指标名称”
    direct = _safe_strip(row.get(COL_INDICATOR_NAME))
    if direct:
        return direct
    # 兜底：处理可能存在的空格差异（例如“子公司 月报表 指标名称”）
    target_key = re.sub(r"\s+", "", COL_INDICATOR_NAME)
    for key, value in row.items():
        key_text = _safe_strip(key)
        if key_text and re.sub(r"\s+", "", key_text) == target_key:
            v = _safe_strip(value)
            if v:
                return v
    return ""


def _build_column_index(header_row: List[Any]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = _safe_strip(cell)
        if name and name not in index:
            index[name] = idx
    return index


def _find_header_row(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    max_scan = min(len(rows), 40)
    for row_idx in range(max_scan):
        col_index = _build_column_index(rows[row_idx])
        if all(col in col_index for col in REQUIRED_COLUMNS):
            return row_idx, col_index
    raise ValueError("映射表未找到必要列：子公司月报表源文件 / 目标文件")


def _read_mapping_rows(mapping_path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    workbook = load_workbook(filename=str(mapping_path), data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        raise ValueError("映射表为空")

    header_row_idx, col_index = _find_header_row(values)
    headers = [_safe_strip(v) for v in values[header_row_idx]]
    rows: List[Dict[str, Any]] = []
    for raw in values[header_row_idx + 1 :]:
        row_dict: Dict[str, Any] = {}
        has_value = False
        for col_name, idx in col_index.items():
            cell_value = raw[idx] if idx < len(raw) else None
            row_dict[col_name] = cell_value
            if _safe_strip(cell_value):
                has_value = True
        if has_value:
            rows.append(row_dict)
    workbook.close()
    return rows, headers


def analyze_mapping(mapping_path: Path) -> Dict[str, Any]:
    rows, columns = _read_mapping_rows(mapping_path)

    adj: Dict[str, Set[str]] = {}
    all_src: Set[str] = set()
    all_tgt: Set[str] = set()
    sheet_requirements: Dict[str, Set[str]] = {}

    for row in rows:
        src_key = _safe_strip(row.get(COL_SRC_KEY))
        tgt_key = _safe_strip(row.get(COL_TGT_KEY))
        if not src_key or not tgt_key:
            continue

        all_src.add(src_key)
        all_tgt.add(tgt_key)
        src_node = f"SRC::{src_key}"
        tgt_node = f"TGT::{tgt_key}"
        adj.setdefault(src_node, set()).add(tgt_node)
        adj.setdefault(tgt_node, set()).add(src_node)

        src_sheet = _safe_strip(row.get(COL_SRC_SHEET))
        if src_sheet:
            sheet_requirements.setdefault(src_key, set()).add(src_sheet)

    groups: List[Dict[str, Any]] = []
    visited: Set[str] = set()
    for node in list(adj.keys()):
        if node in visited:
            continue
        stack = [node]
        visited.add(node)
        group_src: Set[str] = set()
        group_tgt: Set[str] = set()
        while stack:
            current = stack.pop()
            if current.startswith("SRC::"):
                group_src.add(current[5:])
            elif current.startswith("TGT::"):
                group_tgt.add(current[5:])
            for neighbor in adj.get(current, set()):
                if neighbor not in visited:
                    visited.add(neighbor)
                    stack.append(neighbor)
        groups.append(
            {
                "id": len(groups),
                "src": sorted(group_src),
                "tgt": sorted(group_tgt),
            }
        )

    return {
        "columns": columns,
        "groups": groups,
        "required_src_files": sorted(all_src),
        "required_tgt_files": sorted(all_tgt),
        "sheet_requirements": {k: sorted(v) for k, v in sheet_requirements.items()},
        "total_rows": len(rows),
    }


def get_sheet_names(file_path: Path) -> List[str]:
    workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    names = list(workbook.sheetnames)
    workbook.close()
    return names


def _cell_value_as_number(
    sheet,
    cell_ref: str,
    workbook=None,
    default_sheet_name: str = "",
    depth: int = 0,
    visiting: Optional[Set[str]] = None,
) -> float:
    if depth > 8:
        return 0.0
    if visiting is None:
        visiting = set()
    normalized_ref = _normalize_cell_ref(cell_ref)
    visit_key = f"{default_sheet_name}!{normalized_ref}"
    if visit_key in visiting:
        return 0.0
    try:
        value = sheet[normalized_ref].value
    except Exception:
        return 0.0
    if value is None:
        return 0.0
    if isinstance(value, str) and value.strip().startswith("="):
        formula_expr = value.strip()[1:]
        visiting_next = set(visiting)
        visiting_next.add(visit_key)
        evaluated = _evaluate_expr(
            sheet,
            formula_expr,
            workbook=workbook,
            default_sheet_name=default_sheet_name,
            depth=depth + 1,
            visiting=visiting_next,
        )
        return float(evaluated) if evaluated is not None else 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def _normalize_cell_ref(cell_ref: str) -> str:
    return str(cell_ref or "").replace("$", "").upper()


def _sheet_value_by_name(
    workbook,
    sheet_name: str,
    cell_ref: str,
    depth: int = 0,
    visiting: Optional[Set[str]] = None,
) -> float:
    try:
        if sheet_name not in workbook.sheetnames:
            return 0.0
        target_sheet = workbook[sheet_name]
        return _cell_value_as_number(
            target_sheet,
            cell_ref,
            workbook=workbook,
            default_sheet_name=sheet_name,
            depth=depth,
            visiting=visiting,
        )
    except Exception:
        return 0.0


def _evaluate_expr(
    sheet,
    expr: str,
    workbook=None,
    default_sheet_name: str = "",
    depth: int = 0,
    visiting: Optional[Set[str]] = None,
) -> Optional[float]:
    text = _safe_strip(expr)
    if not text:
        return None
    replaced = text

    def _replace_sheet_ref(match: re.Match) -> str:
        sheet_name = _safe_strip(match.group(1) or match.group(2))
        ref = _normalize_cell_ref(f"{match.group(3)}{match.group(4)}")
        if workbook is None:
            return "0.0"
        return str(_sheet_value_by_name(workbook, sheet_name, ref, depth=depth, visiting=visiting))

    replaced = SHEET_CELL_REF_PATTERN.sub(_replace_sheet_ref, replaced)

    refs = set(CELL_REF_PATTERN.findall(replaced))
    for ref in sorted(refs, key=len, reverse=True):
        normalized_ref = _normalize_cell_ref(ref)
        replaced = re.sub(
            rf"\b{ref}\b",
            str(
                _cell_value_as_number(
                    sheet,
                    normalized_ref,
                    workbook=workbook,
                    default_sheet_name=default_sheet_name,
                    depth=depth,
                    visiting=visiting,
                )
            ),
            replaced,
        )

    if not set(replaced).issubset(set("0123456789.+-*/() ")):
        return None
    try:
        return float(eval(replaced))
    except Exception:
        return None


def _extract_cell_refs(expr: str) -> List[str]:
    text = _safe_strip(expr)
    if not text:
        return []
    refs: Set[str] = set()
    for m in SHEET_CELL_REF_PATTERN.finditer(text):
        sheet_name = _safe_strip(m.group(1) or m.group(2))
        cell_ref = _normalize_cell_ref(f"{m.group(3)}{m.group(4)}")
        refs.add(f"{sheet_name}!{cell_ref}")
    cleaned = SHEET_CELL_REF_PATTERN.sub(" ", text)
    for ref in CELL_REF_PATTERN.findall(cleaned):
        refs.add(_normalize_cell_ref(ref))
    return sorted(refs)


def _is_empty_cell_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _collect_empty_refs(sheet, expr: str, workbook=None, default_sheet_name: str = "") -> List[str]:
    refs = _extract_cell_refs(expr)
    empty_refs: List[str] = []
    for ref in refs:
        target_sheet = sheet
        target_ref = ref
        if "!" in ref:
            sheet_name, target_ref = ref.split("!", 1)
            if workbook is None or sheet_name not in getattr(workbook, "sheetnames", []):
                cell_val = None
                if _is_empty_cell_value(cell_val):
                    empty_refs.append(ref)
                continue
            target_sheet = workbook[sheet_name]
        try:
            cell_val = target_sheet[target_ref].value
        except Exception:
            cell_val = None
        if _is_empty_cell_value(cell_val):
            empty_refs.append(ref)
    return empty_refs


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _resolve_source_sheet(workbook, rule_sheet: str, file_sheet_value: str) -> str:
    sheet_map: Dict[str, str] = {}
    default_sheet = file_sheet_value
    raw = _safe_strip(file_sheet_value)
    if raw.startswith("{"):
        try:
            sheet_map = json.loads(raw)
        except Exception:
            sheet_map = {}
        if sheet_map:
            default_sheet = str(next(iter(sheet_map.values())))

    rule_sheet_name = _safe_strip(rule_sheet)
    if rule_sheet_name and rule_sheet_name in sheet_map:
        return _safe_strip(sheet_map[rule_sheet_name]) or default_sheet

    if rule_sheet_name and rule_sheet_name in workbook.sheetnames:
        return rule_sheet_name
    return default_sheet if default_sheet in workbook.sheetnames else workbook.sheetnames[0]


def execute_mapping(
    mapping_path: Path,
    src_files: Dict[str, StoredFileInfo],
    tgt_files: Dict[str, StoredFileInfo],
    bucket_paths: Dict[str, Path],
) -> List[Path]:
    rows, _ = _read_mapping_rows(mapping_path)
    outputs_dir = bucket_paths["outputs"]
    outputs_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    execution_logs: List[Dict[str, Any]] = []

    out_targets: Dict[str, Path] = {}
    for key, info in tgt_files.items():
        source_path = bucket_paths[info.bucket] / info.stored_name
        if not source_path.exists():
            raise FileNotFoundError(f"目标文件不存在：{source_path}")
        clean_stem = source_path.stem
        out_path = outputs_dir / f"{clean_stem}_{timestamp}{source_path.suffix or '.xlsx'}"
        shutil.copy2(source_path, out_path)
        out_targets[key] = out_path

    src_workbooks = {}
    tgt_workbooks = {}
    try:
        for _, info in src_files.items():
            path = (bucket_paths[info.bucket] / info.stored_name).resolve()
            if path not in src_workbooks:
                # 源文件读取计算值，避免将公式字符串直接写入目标后触发 #REF!
                src_workbooks[path] = load_workbook(filename=str(path), data_only=True)
        for path in out_targets.values():
            if path not in tgt_workbooks:
                tgt_workbooks[path] = load_workbook(filename=str(path), data_only=False)

        # 映射文件第一行为表头，异常提示按用户可见行号显示（数据行从第2行开始）
        for idx, row in enumerate(rows, start=2):
            src_key = _safe_strip(row.get(COL_SRC_KEY))
            tgt_key = _safe_strip(row.get(COL_TGT_KEY))
            if src_key not in src_files or tgt_key not in tgt_files:
                execution_logs.append(
                    {
                        "row_index": idx,
                        "status": "skip_missing_key",
                        "src_key": src_key,
                        "tgt_key": tgt_key,
                    }
                )
                continue

            src_info = src_files[src_key]
            tgt_info = tgt_files[tgt_key]
            src_path = (bucket_paths[src_info.bucket] / src_info.stored_name).resolve()
            tgt_out_path = out_targets[tgt_key]

            src_wb = src_workbooks[src_path]
            tgt_wb = tgt_workbooks[tgt_out_path]

            src_sheet_name = _resolve_source_sheet(src_wb, _safe_strip(row.get(COL_SRC_SHEET)), src_info.sheet)
            tgt_sheet_name = _safe_strip(tgt_info.sheet)
            if tgt_sheet_name not in tgt_wb.sheetnames:
                tgt_sheet_name = tgt_wb.sheetnames[0]

            src_sheet = src_wb[src_sheet_name]
            tgt_sheet = tgt_wb[tgt_sheet_name]

            src_cell_expr = _safe_strip(row.get(COL_SRC_CELL))
            tgt_cell = _safe_strip(row.get(COL_TGT_CELL))
            row_log: Dict[str, Any] = {
                "row_index": idx,
                "src_key": src_key,
                "tgt_key": tgt_key,
                "indicator_name": _extract_indicator_name(row),
                "src_sheet": src_sheet_name,
                "tgt_sheet": tgt_sheet_name,
                "src_cell_expr": src_cell_expr,
                "tgt_cell": tgt_cell,
                "status": "ok",
            }

            month_value = None
            month_empty_refs: List[str] = []
            if src_cell_expr:
                if SIMPLE_CELL_PATTERN.fullmatch(src_cell_expr):
                    month_value = src_sheet[src_cell_expr].value
                    if _is_empty_cell_value(month_value):
                        month_empty_refs.append(src_cell_expr)
                else:
                    month_value = _evaluate_expr(
                        src_sheet,
                        src_cell_expr,
                        workbook=src_wb,
                        default_sheet_name=src_sheet_name,
                    )
                    month_empty_refs = _collect_empty_refs(
                        src_sheet,
                        src_cell_expr,
                        workbook=src_wb,
                        default_sheet_name=src_sheet_name,
                    )
                    if month_value is None:
                        row_log["status"] = "warn_month_expr_invalid"
                        row_log["message"] = f"源本月表达式无法计算：{src_cell_expr}"
            if month_value is not None and tgt_cell:
                try:
                    tgt_sheet[tgt_cell].value = month_value
                    row_log["month_value"] = month_value
                except Exception as exc:
                    row_log["status"] = "error_write_month"
                    row_log["month_value"] = month_value
                    row_log["error"] = str(exc)

            action = _safe_strip(row.get(COL_ACTION))
            if "粘贴" in action and "累计" in action:
                src_acc = _safe_strip(row.get(COL_SRC_ACC))
                tgt_acc = _safe_strip(row.get(COL_TGT_ACC))
                row_log["src_acc"] = src_acc
                row_log["tgt_acc"] = tgt_acc
                if src_acc and tgt_acc:
                    try:
                        if SIMPLE_CELL_PATTERN.fullmatch(src_acc):
                            acc_val = src_sheet[src_acc].value
                            acc_empty_refs = [src_acc] if _is_empty_cell_value(acc_val) else []
                        else:
                            acc_val = _evaluate_expr(
                                src_sheet,
                                src_acc,
                                workbook=src_wb,
                                default_sheet_name=src_sheet_name,
                            )
                            acc_empty_refs = _collect_empty_refs(
                                src_sheet,
                                src_acc,
                                workbook=src_wb,
                                default_sheet_name=src_sheet_name,
                            )
                            if acc_val is None:
                                row_log["status"] = "warn_acc_expr_invalid"
                                row_log["message"] = f"源累计表达式无法计算：{src_acc}"
                        row_log["acc_value"] = acc_val
                        existing_before = tgt_sheet[tgt_acc].value
                        row_log["tgt_acc_before"] = existing_before
                        if acc_val is not None:
                            existing = existing_before
                            # 目标累计单元格若已有公式，保持公式不被覆盖
                            if isinstance(existing, str) and existing.strip().startswith("="):
                                row_log["acc_action"] = "preserve_formula"
                                formula_expr = existing.strip()[1:]
                                row_log["tgt_acc_formula"] = existing
                                evaluated_target = _evaluate_expr(
                                    tgt_sheet,
                                    formula_expr,
                                    workbook=tgt_wb,
                                    default_sheet_name=tgt_sheet_name,
                                )
                                src_num = _to_float(acc_val)
                                tgt_num = _to_float(evaluated_target)
                                if src_num is None or tgt_num is None:
                                    row_log["acc_compare_status"] = "formula_not_verifiable"
                                else:
                                    diff = round(tgt_num - src_num, 10)
                                    row_log["acc_compare_status"] = "ok" if abs(diff) < 1e-9 else "mismatch"
                                    row_log["acc_compare_diff"] = diff
                            else:
                                tgt_sheet[tgt_acc].value = acc_val
                                row_log["acc_action"] = "write_value"
                                after_val = tgt_sheet[tgt_acc].value
                                src_num = _to_float(acc_val)
                                tgt_num = _to_float(after_val)
                                if src_num is None or tgt_num is None:
                                    row_log["acc_compare_status"] = "non_numeric"
                                else:
                                    diff = round(tgt_num - src_num, 10)
                                    row_log["acc_compare_status"] = "ok" if abs(diff) < 1e-9 else "mismatch"
                                    row_log["acc_compare_diff"] = diff
                        if acc_empty_refs:
                            row_log["empty_source_refs_acc"] = acc_empty_refs
                    except Exception as exc:
                        row_log["status"] = "error_write_acc"
                        row_log["error"] = str(exc)

            if isinstance(month_value, str) and month_value.strip().startswith("="):
                row_log["status"] = "warn_formula_text_from_source"
                row_log["message"] = "源单元格读取到公式文本，可能导致目标出现公式引用问题。"
            if month_empty_refs:
                row_log["empty_source_refs_month"] = month_empty_refs
            all_empty_refs = list({*row_log.get("empty_source_refs_month", []), *row_log.get("empty_source_refs_acc", [])})
            if all_empty_refs and not str(row_log.get("status", "")).startswith("error_"):
                row_log["status"] = "warn_source_empty"
                row_log["message"] = f"源单元格为空：{', '.join(all_empty_refs)}"
            execution_logs.append(row_log)

        final_outputs: List[Path] = []
        for path, workbook in tgt_workbooks.items():
            workbook.save(str(path))
            final_outputs.append(path)
        compare_stats = {"ok": 0, "mismatch": 0, "skipped_target_formula": 0, "non_numeric": 0, "formula_not_verifiable": 0}
        for item in execution_logs:
            status = str(item.get("acc_compare_status") or "")
            if status in compare_stats:
                compare_stats[status] += 1
        log_payload = {
            "timestamp": timestamp,
            "mapping_file": str(mapping_path.name),
            "rows_total": len(rows),
            "acc_compare_stats": compare_stats,
            "logs": execution_logs,
        }
        log_path = outputs_dir / f"execution_log_{timestamp}.json"
        log_path.write_text(json.dumps(log_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        final_outputs.append(log_path)
        return final_outputs
    finally:
        for workbook in src_workbooks.values():
            workbook.close()
        for workbook in tgt_workbooks.values():
            workbook.close()
