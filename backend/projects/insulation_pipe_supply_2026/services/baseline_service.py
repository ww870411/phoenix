# -*- coding: utf-8 -*-
"""
保温直管与管件及标准化物料设计量与计划采购量基准服务 (Baseline Service)。
支持数据库表自愈、多维查询、批量保存以及从 Excel / JSON 平滑导入入库。
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional
import openpyxl
from sqlalchemy import text

from backend.db.database_daily_report_25_26 import SessionLocal


_baseline_tables_checked = False


def ensure_baseline_tables() -> None:
    """自愈检查并创建/升级 tube_pipe_baseline 与 tube_fitting_baseline 表及索引。"""
    global _baseline_tables_checked
    if _baseline_tables_checked:
        return

    session = SessionLocal()
    try:
        # 1. 确保 schema 存在
        session.execute(text("CREATE SCHEMA IF NOT EXISTS tube;"))

        # 2. 直管基准表
        session.execute(text("""
            CREATE TABLE IF NOT EXISTS tube.tube_pipe_baseline (
                id BIGSERIAL PRIMARY KEY,
                section_1_id VARCHAR(64) NOT NULL,
                pipe_model_id VARCHAR(128) NOT NULL,
                unit VARCHAR(32) NOT NULL DEFAULT '米',
                design_qty NUMERIC(18, 2) NOT NULL DEFAULT 0,
                purchase_plan_qty NUMERIC(18, 2) NOT NULL DEFAULT 0,
                remark TEXT,
                created_by VARCHAR(128),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_by VARCHAR(128),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                CONSTRAINT chk_tube_pipe_baseline_qty_nonnegative 
                    CHECK (design_qty >= 0 AND purchase_plan_qty >= 0)
            );
        """))
        session.execute(text("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_tube_pipe_baseline_sec_model 
                ON tube.tube_pipe_baseline (section_1_id, pipe_model_id);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_pipe_baseline_sec 
                ON tube.tube_pipe_baseline (section_1_id);
        """))

        # 确保直管序列与默认值
        try:
            session.execute(text("""
                CREATE SEQUENCE IF NOT EXISTS tube.tube_pipe_baseline_id_seq;
                ALTER TABLE tube.tube_pipe_baseline ALTER COLUMN id SET DEFAULT nextval('tube.tube_pipe_baseline_id_seq');
            """))
            session.commit()
        except Exception:
            session.rollback()

        # 3. 管件与标准化物料基准表 (自愈升级)
        session.execute(text("""
            CREATE TABLE IF NOT EXISTS tube.tube_fitting_baseline (
                id BIGSERIAL PRIMARY KEY,
                section_1_id VARCHAR(64) NOT NULL,
                system_type VARCHAR(32) NOT NULL DEFAULT '高温水',
                category VARCHAR(64) NOT NULL DEFAULT '管件',
                standard_name VARCHAR(128) NOT NULL DEFAULT '',
                model_spec VARCHAR(255) NOT NULL,
                sub_model_spec VARCHAR(128) NOT NULL DEFAULT '',
                unit VARCHAR(32) NOT NULL DEFAULT '个',
                design_qty NUMERIC(18, 2) NOT NULL DEFAULT 0,
                purchase_plan_qty NUMERIC(18, 2) NOT NULL DEFAULT 0,
                main_dn NUMERIC(10, 2),
                sub_dn NUMERIC(10, 2),
                angle NUMERIC(10, 2),
                bending_radius_ratio NUMERIC(10, 2),
                bending_radius_m NUMERIC(10, 2),
                valve_model VARCHAR(128),
                outer_diameter NUMERIC(10, 2),
                wall_thickness NUMERIC(10, 2),
                length_m NUMERIC(10, 2),
                pressure_rating VARCHAR(64),
                compensation_mm NUMERIC(10, 2),
                flow_direction VARCHAR(64),
                raw_model_spec VARCHAR(255),
                raw_name VARCHAR(128),
                remark TEXT,
                extra_params JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_by VARCHAR(128),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_by VARCHAR(128),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                CONSTRAINT chk_tube_fitting_baseline_qty_nonnegative 
                    CHECK (design_qty >= 0 AND purchase_plan_qty >= 0)
            );
        """))

        # 补齐可能缺失的字段 (向后自愈)
        alter_cols = [
            ("system_type", "VARCHAR(32) NOT NULL DEFAULT '高温水'"),
            ("category", "VARCHAR(64) NOT NULL DEFAULT '管件'"),
            ("standard_name", "VARCHAR(128) NOT NULL DEFAULT ''"),
            ("sub_model_spec", "VARCHAR(128) NOT NULL DEFAULT ''"),
            ("main_dn", "NUMERIC(10, 2)"),
            ("sub_dn", "NUMERIC(10, 2)"),
            ("angle", "NUMERIC(10, 2)"),
            ("bending_radius_ratio", "NUMERIC(10, 2)"),
            ("bending_radius_m", "NUMERIC(10, 2)"),
            ("valve_model", "VARCHAR(128)"),
            ("outer_diameter", "NUMERIC(10, 2)"),
            ("wall_thickness", "NUMERIC(10, 2)"),
            ("length_m", "NUMERIC(10, 2)"),
            ("pressure_rating", "VARCHAR(64)"),
            ("compensation_mm", "NUMERIC(10, 2)"),
            ("flow_direction", "VARCHAR(64)"),
            ("raw_model_spec", "VARCHAR(255)"),
            ("raw_name", "VARCHAR(128)"),
            ("extra_params", "JSONB NOT NULL DEFAULT '{}'::jsonb"),
        ]
        for col_name, col_def in alter_cols:
            try:
                session.execute(text(f"ALTER TABLE tube.tube_fitting_baseline ADD COLUMN IF NOT EXISTS {col_name} {col_def};"))
            except Exception:
                pass

        # 适配旧表 fitting_type -> category 数据迁移 (如果存在 fitting_type 字段)
        try:
            session.execute(text("""
                DO $$
                BEGIN
                    IF EXISTS (
                        SELECT 1 FROM information_schema.columns 
                        WHERE table_schema = 'tube' AND table_name = 'tube_fitting_baseline' AND column_name = 'fitting_type'
                    ) THEN
                        UPDATE tube.tube_fitting_baseline 
                        SET category = fitting_type 
                        WHERE category = '管件' OR category IS NULL OR category = '';
                    END IF;
                END $$;
            """))
        except Exception:
            pass

        # 删除旧唯一索引（如果有）
        try:
            session.execute(text("DROP INDEX IF EXISTS tube.uq_tube_fitting_baseline_sec_type_spec_sub;"))
        except Exception:
            pass

        # 创建新联合唯一索引与高频查询索引
        session.execute(text("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_tube_fitting_baseline_sec_sys_name_spec_sub 
                ON tube.tube_fitting_baseline (section_1_id, system_type, standard_name, model_spec, sub_model_spec);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_fitting_baseline_sec_sys 
                ON tube.tube_fitting_baseline (section_1_id, system_type);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_fitting_baseline_category 
                ON tube.tube_fitting_baseline (category);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_fitting_baseline_standard_name 
                ON tube.tube_fitting_baseline (standard_name);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_fitting_baseline_main_dn 
                ON tube.tube_fitting_baseline (main_dn);
        """))

        # 序列自愈校准
        try:
            session.execute(text("""
                CREATE SEQUENCE IF NOT EXISTS tube.tube_pipe_baseline_id_seq;
                ALTER TABLE tube.tube_pipe_baseline ALTER COLUMN id SET DEFAULT nextval('tube.tube_pipe_baseline_id_seq');
                SELECT setval('tube.tube_pipe_baseline_id_seq', COALESCE((SELECT MAX(id) FROM tube.tube_pipe_baseline), 0) + 1, false);
                CREATE SEQUENCE IF NOT EXISTS tube.tube_fitting_baseline_id_seq;
                ALTER TABLE tube.tube_fitting_baseline ALTER COLUMN id SET DEFAULT nextval('tube.tube_fitting_baseline_id_seq');
                SELECT setval('tube.tube_fitting_baseline_id_seq', COALESCE((SELECT MAX(id) FROM tube.tube_fitting_baseline), 0) + 1, false);
            """))
        except Exception:
            pass

        session.commit()
        _baseline_tables_checked = True
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _clean_str(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" or s.lower() == "none" else s


def _clean_num(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        num = float(val)
        return num if num >= 0 else 0.0
    except (ValueError, TypeError):
        return 0.0


def _clean_num_nullable(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        s = str(val).strip()
        if s.lower() in ("nan", "none", ""):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


# -----------------------------------------------------------------------------
# 直管基准量 (Pipe Baseline) 核心操作
# -----------------------------------------------------------------------------

def list_pipe_baselines(section_1_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """查询直管基准列表。"""
    ensure_baseline_tables()
    session = SessionLocal()
    try:
        sql = """
            SELECT id, section_1_id, pipe_model_id, unit, design_qty, purchase_plan_qty,
                   remark, created_by, created_at, updated_by, updated_at
            FROM tube.tube_pipe_baseline
        """
        params = {}
        if section_1_id:
            sql += " WHERE section_1_id = :section_1_id"
            params["section_1_id"] = section_1_id
        sql += " ORDER BY section_1_id ASC, id ASC"

        rows = session.execute(text(sql), params).mappings().all()
        return [
            {
                "id": row["id"],
                "section_1_id": row["section_1_id"],
                "pipe_model_id": row["pipe_model_id"],
                "unit": row["unit"] or "米",
                "design_qty": float(row["design_qty"] or 0),
                "purchase_plan_qty": float(row["purchase_plan_qty"] or 0),
                "remark": row["remark"] or "",
                "created_by": row["created_by"] or "",
                "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                "updated_by": row["updated_by"] or "",
                "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            }
            for row in rows
        ]
    finally:
        session.close()


def save_pipe_baselines(
    items: List[Dict[str, Any]], 
    operator_name: str = "",
    replace_all_for_sections: Optional[List[str]] = None
) -> int:
    """保存或更新直管基准记录。"""
    ensure_baseline_tables()
    session = SessionLocal()
    try:
        if replace_all_for_sections:
            session.execute(
                text("DELETE FROM tube.tube_pipe_baseline WHERE section_1_id = ANY(:sec_list)"),
                {"sec_list": list(set(replace_all_for_sections))}
            )

        upsert_sql = """
            INSERT INTO tube.tube_pipe_baseline (
                section_1_id, pipe_model_id, unit, design_qty, purchase_plan_qty,
                remark, created_by, updated_by, updated_at
            ) VALUES (
                :section_1_id, :pipe_model_id, :unit, :design_qty, :purchase_plan_qty,
                :remark, :operator_name, :operator_name, NOW()
            )
            ON CONFLICT (section_1_id, pipe_model_id) DO UPDATE SET
                unit = EXCLUDED.unit,
                design_qty = EXCLUDED.design_qty,
                purchase_plan_qty = EXCLUDED.purchase_plan_qty,
                remark = EXCLUDED.remark,
                updated_by = EXCLUDED.updated_by,
                updated_at = NOW()
        """
        saved_count = 0
        for it in items:
            sec_id = _clean_str(it.get("section_1_id"))
            model_id = _clean_str(it.get("pipe_model_id"))
            if not sec_id or not model_id:
                continue

            session.execute(
                text(upsert_sql),
                {
                    "section_1_id": sec_id,
                    "pipe_model_id": model_id,
                    "unit": _clean_str(it.get("unit")) or "米",
                    "design_qty": _clean_num(it.get("design_qty")),
                    "purchase_plan_qty": _clean_num(it.get("purchase_plan_qty")),
                    "remark": _clean_str(it.get("remark")),
                    "operator_name": _clean_str(operator_name) or "system",
                }
            )
            saved_count += 1

        session.commit()
        return saved_count
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


# -----------------------------------------------------------------------------
# 管件与物料基准量 (Fitting Baseline) 核心操作
# -----------------------------------------------------------------------------

def list_fitting_baselines(
    section_1_id: Optional[str] = None,
    system_type: Optional[str] = None,
    category: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """查询多维管件与物料基准列表。"""
    ensure_baseline_tables()
    session = SessionLocal()
    try:
        sql = """
            SELECT id, section_1_id, system_type, category, standard_name, model_spec, sub_model_spec,
                   unit, design_qty, purchase_plan_qty,
                   main_dn, sub_dn, angle, bending_radius_ratio, bending_radius_m,
                   valve_model, outer_diameter, wall_thickness, length_m,
                   pressure_rating, compensation_mm, flow_direction,
                   raw_model_spec, raw_name, remark, extra_params,
                   created_by, created_at, updated_by, updated_at
            FROM tube.tube_fitting_baseline
            WHERE 1=1
        """
        params: Dict[str, Any] = {}
        if section_1_id:
            sql += " AND section_1_id = :section_1_id"
            params["section_1_id"] = section_1_id
        if system_type:
            sql += " AND system_type = :system_type"
            params["system_type"] = system_type
        if category:
            sql += " AND category = :category"
            params["category"] = category

        sql += " ORDER BY section_1_id ASC, system_type ASC, category ASC, main_dn DESC NULLS LAST, id ASC"

        rows = session.execute(text(sql), params).mappings().all()
        return [
            {
                "id": row["id"],
                "section_1_id": row["section_1_id"],
                "system_type": row["system_type"] or "高温水",
                "category": row["category"] or "管件",
                "fitting_type": row["category"] or "管件",  # 向后兼容
                "standard_name": row["standard_name"] or "",
                "model_spec": row["model_spec"],
                "sub_model_spec": row["sub_model_spec"] or "",
                "unit": row["unit"] or "个",
                "design_qty": float(row["design_qty"] or 0),
                "purchase_plan_qty": float(row["purchase_plan_qty"] or 0),
                "main_dn": float(row["main_dn"]) if row["main_dn"] is not None else None,
                "sub_dn": float(row["sub_dn"]) if row["sub_dn"] is not None else None,
                "angle": float(row["angle"]) if row["angle"] is not None else None,
                "bending_radius_ratio": float(row["bending_radius_ratio"]) if row["bending_radius_ratio"] is not None else None,
                "bending_radius_m": float(row["bending_radius_m"]) if row["bending_radius_m"] is not None else None,
                "valve_model": row["valve_model"] or "",
                "outer_diameter": float(row["outer_diameter"]) if row["outer_diameter"] is not None else None,
                "wall_thickness": float(row["wall_thickness"]) if row["wall_thickness"] is not None else None,
                "length_m": float(row["length_m"]) if row["length_m"] is not None else None,
                "pressure_rating": row["pressure_rating"] or "",
                "compensation_mm": float(row["compensation_mm"]) if row["compensation_mm"] is not None else None,
                "flow_direction": row["flow_direction"] or "",
                "raw_model_spec": row.get("raw_model_spec") or "",
                "raw_name": row.get("raw_name") or "",
                "remark": row["remark"] or "",
                "extra_params": row["extra_params"] or {},
                "created_by": row["created_by"] or "",
                "created_at": row["created_at"].isoformat() if row["created_at"] else None,
                "updated_by": row["updated_by"] or "",
                "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
            }
            for row in rows
        ]
    finally:
        session.close()


def save_fitting_baselines(
    items: List[Dict[str, Any]], 
    operator_name: str = "",
    replace_all_for_sections: Optional[List[str]] = None
) -> int:
    """保存或更新管件与多维物料基准记录 (支持全量 UPSERT)。"""
    ensure_baseline_tables()
    session = SessionLocal()
    try:
        if replace_all_for_sections:
            session.execute(
                text("DELETE FROM tube.tube_fitting_baseline WHERE section_1_id = ANY(:sec_list)"),
                {"sec_list": list(set(replace_all_for_sections))}
            )

        upsert_sql = """
            INSERT INTO tube.tube_fitting_baseline (
                section_1_id, system_type, category, standard_name, model_spec, sub_model_spec, unit,
                design_qty, purchase_plan_qty,
                main_dn, sub_dn, angle, bending_radius_ratio, bending_radius_m,
                valve_model, outer_diameter, wall_thickness, length_m,
                pressure_rating, compensation_mm, flow_direction, raw_model_spec, raw_name, remark, extra_params,
                created_by, updated_by, updated_at
            ) VALUES (
                :section_1_id, :system_type, :category, :standard_name, :model_spec, :sub_model_spec, :unit,
                :design_qty, :purchase_plan_qty,
                :main_dn, :sub_dn, :angle, :bending_radius_ratio, :bending_radius_m,
                :valve_model, :outer_diameter, :wall_thickness, :length_m,
                :pressure_rating, :compensation_mm, :flow_direction, :raw_model_spec, :raw_name, :remark, '{}'::jsonb,
                :operator_name, :operator_name, NOW()
            )
            ON CONFLICT (section_1_id, system_type, standard_name, model_spec, sub_model_spec) DO UPDATE SET
                category = EXCLUDED.category,
                unit = EXCLUDED.unit,
                design_qty = EXCLUDED.design_qty,
                purchase_plan_qty = EXCLUDED.purchase_plan_qty,
                main_dn = EXCLUDED.main_dn,
                sub_dn = EXCLUDED.sub_dn,
                angle = EXCLUDED.angle,
                bending_radius_ratio = EXCLUDED.bending_radius_ratio,
                bending_radius_m = EXCLUDED.bending_radius_m,
                valve_model = EXCLUDED.valve_model,
                outer_diameter = EXCLUDED.outer_diameter,
                wall_thickness = EXCLUDED.wall_thickness,
                length_m = EXCLUDED.length_m,
                pressure_rating = EXCLUDED.pressure_rating,
                compensation_mm = EXCLUDED.compensation_mm,
                flow_direction = EXCLUDED.flow_direction,
                raw_model_spec = EXCLUDED.raw_model_spec,
                raw_name = EXCLUDED.raw_name,
                remark = EXCLUDED.remark,
                updated_by = EXCLUDED.updated_by,
                updated_at = NOW()
        """
        saved_count = 0
        for it in items:
            sec_id = _clean_str(it.get("section_1_id"))
            sys_type = _clean_str(it.get("system_type")) or "高温水"
            category = _clean_str(it.get("category") or it.get("fitting_type")) or "管件"
            std_name = _clean_str(it.get("standard_name"))
            m_spec = _clean_str(it.get("model_spec"))
            sub_spec = _clean_str(it.get("sub_model_spec"))
            
            if not sec_id or not m_spec:
                continue

            session.execute(
                text(upsert_sql),
                {
                    "section_1_id": sec_id,
                    "system_type": sys_type,
                    "category": category,
                    "standard_name": std_name,
                    "model_spec": m_spec,
                    "sub_model_spec": sub_spec,
                    "unit": _clean_str(it.get("unit")) or "个",
                    "design_qty": _clean_num(it.get("design_qty")),
                    "purchase_plan_qty": _clean_num(it.get("purchase_plan_qty")),
                    "main_dn": _clean_num_nullable(it.get("main_dn")),
                    "sub_dn": _clean_num_nullable(it.get("sub_dn")),
                    "angle": _clean_num_nullable(it.get("angle")),
                    "bending_radius_ratio": _clean_num_nullable(it.get("bending_radius_ratio")),
                    "bending_radius_m": _clean_num_nullable(it.get("bending_radius_m")),
                    "valve_model": _clean_str(it.get("valve_model")),
                    "outer_diameter": _clean_num_nullable(it.get("outer_diameter")),
                    "wall_thickness": _clean_num_nullable(it.get("wall_thickness")),
                    "length_m": _clean_num_nullable(it.get("length_m")),
                    "pressure_rating": _clean_str(it.get("pressure_rating")),
                    "compensation_mm": _clean_num_nullable(it.get("compensation_mm")),
                    "flow_direction": _clean_str(it.get("flow_direction")),
                    "raw_model_spec": _clean_str(it.get("raw_model_spec")),
                    "raw_name": _clean_str(it.get("raw_name")),
                    "remark": _clean_str(it.get("remark")),
                    "operator_name": _clean_str(operator_name) or "system",
                }
            )
            saved_count += 1

        session.commit()
        return saved_count
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def import_fitting_baselines_from_excel(
    excel_path: str,
    operator_name: str = "excel_import",
    replace_all: bool = False
) -> Dict[str, Any]:
    """
    从标准化 Excel 文件导入全量管件与物料基准量。
    支持 22 个维度字段全自动识别与映射。
    """
    ensure_baseline_tables()
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel 文件未找到: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = "标准化数据" if "标准化数据" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]

    headers: List[str] = [str(sheet.cell(row=1, column=c).value or "").strip() for c in range(1, sheet.max_column + 1)]
    col_map = {name: idx + 1 for idx, name in enumerate(headers) if name}

    def _get_val(row_idx: int, col_name: str) -> Any:
        c_idx = col_map.get(col_name)
        if not c_idx:
            return None
        return sheet.cell(row=row_idx, column=c_idx).value

    items: List[Dict[str, Any]] = []
    for r in range(2, sheet.max_row + 1):
        sec_id = _clean_str(_get_val(r, "标段ID"))
        m_spec = _clean_str(_get_val(r, "型号规格"))
        if not sec_id or not m_spec:
            continue

        item = {
            "section_1_id": sec_id,
            "system_type": _clean_str(_get_val(r, "系统类型")) or "高温水",
            "category": _clean_str(_get_val(r, "物理类别")) or "管件",
            "standard_name": _clean_str(_get_val(r, "标准名称")),
            "model_spec": m_spec,
            "sub_model_spec": _clean_str(_get_val(r, "子型号规格") or _get_val(r, "子型号")),
            "unit": _clean_str(_get_val(r, "单位")) or "个",
            "design_qty": _clean_num(_get_val(r, "设计使用量")),
            "purchase_plan_qty": _clean_num(_get_val(r, "计划采购量")),
            "main_dn": _clean_num_nullable(_get_val(r, "主径DN")),
            "sub_dn": _clean_num_nullable(_get_val(r, "次径DN")),
            "angle": _clean_num_nullable(_get_val(r, "角度(°)") or _get_val(r, "角度")),
            "bending_radius_ratio": _clean_num_nullable(_get_val(r, "弯曲半径倍数")),
            "bending_radius_m": _clean_num_nullable(_get_val(r, "弯曲半径(m)") or _get_val(r, "弯曲半径")),
            "valve_model": _clean_str(_get_val(r, "阀门型号")),
            "outer_diameter": _clean_num_nullable(_get_val(r, "外径Φ(mm)") or _get_val(r, "外径")),
            "wall_thickness": _clean_num_nullable(_get_val(r, "壁厚(mm)") or _get_val(r, "壁厚")),
            "length_m": _clean_num_nullable(_get_val(r, "长度(m)") or _get_val(r, "长度")),
            "pressure_rating": _clean_str(_get_val(r, "公称压力/压力等级") or _get_val(r, "公称压力")),
            "compensation_mm": _clean_num_nullable(_get_val(r, "补偿量(mm)") or _get_val(r, "补偿量")),
            "flow_direction": _clean_str(_get_val(r, "流向/方向") or _get_val(r, "流向")),
            "raw_model_spec": _clean_str(_get_val(r, "原型号规格") or _get_val(r, "原始型号规格")),
            "raw_name": _clean_str(_get_val(r, "原名称") or _get_val(r, "原始名称")),
            "remark": _clean_str(_get_val(r, "备注")),
        }
        items.append(item)

    replace_sections = list({it["section_1_id"] for it in items}) if replace_all else None
    saved_count = save_fitting_baselines(items, operator_name=operator_name, replace_all_for_sections=replace_sections)

    total_in_db = len(list_fitting_baselines())
    return {
        "ok": True,
        "excel_rows_parsed": len(items),
        "saved_count": saved_count,
        "total_db_count": total_in_db,
    }

