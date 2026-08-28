# -*- coding: utf-8 -*-
"""
保温管与管件物料合同单价与造价基准服务 (Price Service)。
支持表结构自愈检查、Excel 单价字典逐行全量导入 (允许同名同型号多行不同报价记录并保留原始备注)、多维单价检索以及实时金额测算。
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional
import openpyxl
from sqlalchemy import text

from backend.db.database_daily_report_25_26 import SessionLocal
from backend.projects.insulation_pipe_supply_2026.services.config_service import load_tube_config


_price_table_checked = False


def _normalize_text(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _resolve_supply_entity_id(supplier_name: str, config: Dict[str, Any]) -> str:
    """根据供给方全称在 tube_config.json 的 supply_entities 中匹配缩写 ID。"""
    name_clean = _normalize_text(supplier_name)
    entities = config.get("supply_entities", [])
    
    # 1. 精确匹配 entity_name
    for ent in entities:
        if _normalize_text(ent.get("entity_name")) == name_clean:
            return ent.get("entity_id", "")
            
    # 2. 模糊子串匹配
    for ent in entities:
        ent_name = _normalize_text(ent.get("entity_name"))
        if ent_name and (ent_name in name_clean or name_clean in ent_name):
            return ent.get("entity_id", "")
            
    # 3. 常见关键字硬编码别名
    if "开元" in name_clean:
        return "kaiyuan"
    if "鑫瑞得" in name_clean:
        return "xinruide"
    if "沃圣" in name_clean:
        return "wosheng"
    if "卡尔斯" in name_clean:
        return "kaersi"
    if "三维" in name_clean:
        return "sanwei"
    if "华阳" in name_clean:
        return "huayang"
    if "天地龙" in name_clean:
        return "tiandilong"
    if "泽悦" in name_clean:
        return "zeyue"
    if "保温管厂" in name_clean:
        return "吴近"

    return ""


def ensure_price_table() -> None:
    """自愈检查并创建/升级 tube.tube_material_price 表、自增序列与索引 (无唯一约束，允许同名多行报价)。"""
    global _price_table_checked
    if _price_table_checked:
        return

    session = SessionLocal()
    try:
        session.execute(text("CREATE SCHEMA IF NOT EXISTS tube;"))

        session.execute(text("""
            CREATE TABLE IF NOT EXISTS tube.tube_material_price (
                id BIGSERIAL PRIMARY KEY,
                project_key VARCHAR(64) NOT NULL DEFAULT 'insulation_pipe_supply_2026',
                material_kind VARCHAR(32) NOT NULL,
                supply_entity_id VARCHAR(64),
                supplier_name VARCHAR(128) NOT NULL,
                category VARCHAR(64) NOT NULL,
                material_name VARCHAR(128) NOT NULL,
                model_spec VARCHAR(255) NOT NULL,
                raw_model_spec VARCHAR(128),
                unit VARCHAR(32) NOT NULL DEFAULT '米',
                unit_price NUMERIC(18, 2) NOT NULL DEFAULT 0,
                remark TEXT,
                created_by VARCHAR(128) DEFAULT 'EXCEL_IMPORT',
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_by VARCHAR(128) DEFAULT 'EXCEL_IMPORT',
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                CONSTRAINT chk_tube_material_price_unit_price_nonnegative 
                    CHECK (unit_price >= 0),
                CONSTRAINT chk_tube_material_price_material_kind 
                    CHECK (material_kind IN ('pipe', 'fitting'))
            );
        """))

        # 确保自增序列
        try:
            session.execute(text("""
                CREATE SEQUENCE IF NOT EXISTS tube.tube_material_price_id_seq;
                ALTER TABLE tube.tube_material_price ALTER COLUMN id SET DEFAULT nextval('tube.tube_material_price_id_seq');
            """))
            session.commit()
        except Exception:
            session.rollback()

        # 移除旧的唯一索引 (如果存在)
        try:
            session.execute(text("DROP INDEX IF EXISTS tube.uq_tube_material_price_sup_spec;"))
            session.commit()
        except Exception:
            session.rollback()

        # 普通检索索引
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_material_price_sup_spec 
                ON tube.tube_material_price (supplier_name, model_spec);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_material_price_kind_sup 
                ON tube.tube_material_price (material_kind, supplier_name);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_material_price_category 
                ON tube.tube_material_price (category);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_material_price_entity_id 
                ON tube.tube_material_price (supply_entity_id);
        """))
        session.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_tube_material_price_spec 
                ON tube.tube_material_price (model_spec);
        """))

        # 便捷视图
        session.execute(text("""
            CREATE OR REPLACE VIEW tube.v_tube_pipe_price AS
            SELECT * FROM tube.tube_material_price WHERE material_kind = 'pipe';
        """))
        session.execute(text("""
            CREATE OR REPLACE VIEW tube.v_tube_fitting_price AS
            SELECT * FROM tube.tube_material_price WHERE material_kind = 'fitting';
        """))

        session.commit()
        _price_table_checked = True
    except Exception as e:
        session.rollback()
        raise RuntimeError(f"初始化 tube.tube_material_price 表失败: {e}") from e
    finally:
        session.close()


def import_prices_from_excel(
    pipe_excel_path: Optional[str] = None,
    fitting_excel_path: Optional[str] = None,
    operator: str = "EXCEL_IMPORT"
) -> Dict[str, Any]:
    """
    从标准化 Excel 文件逐行全量导入保温管与管件物料单价字典。
    100% 保留原始 Excel 的全部行与原始备注（即使同厂家同型号存在不同报价，也全部作为独立记录入库）。
    """
    ensure_price_table()
    config = load_tube_config()

    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
    if not pipe_excel_path:
        pipe_excel_path = os.path.join(base_dir, "configs", "8.28 保温管价格_标准化.xlsx")
    if not fitting_excel_path:
        fitting_excel_path = os.path.join(base_dir, "configs", "8.28 管件价格_标准化.xlsx")

    all_rows: List[Dict[str, Any]] = []

    # 1. 解析保温直管单价 Excel (逐行读取，保留原始备注)
    if os.path.exists(pipe_excel_path):
        wb_pipe = openpyxl.load_workbook(pipe_excel_path, data_only=True)
        ws_pipe = wb_pipe.active
        for r in range(2, ws_pipe.max_row + 1):
            sup_name = _normalize_text(ws_pipe.cell(r, 1).value)
            category = _normalize_text(ws_pipe.cell(r, 2).value) or "保温管"
            mat_name = _normalize_text(ws_pipe.cell(r, 3).value) or "塑套钢直埋预制保温管"
            model_spec = _normalize_text(ws_pipe.cell(r, 4).value)
            unit = _normalize_text(ws_pipe.cell(r, 5).value) or "米"
            raw_price = ws_pipe.cell(r, 7).value
            remark = _normalize_text(ws_pipe.cell(r, 8).value)

            if not sup_name or not model_spec:
                continue

            try:
                unit_price = round(float(raw_price or 0), 2)
            except (ValueError, TypeError):
                unit_price = 0.0

            entity_id = _resolve_supply_entity_id(sup_name, config)
            all_rows.append({
                "project_key": "insulation_pipe_supply_2026",
                "material_kind": "pipe",
                "supply_entity_id": entity_id,
                "supplier_name": sup_name,
                "category": category,
                "material_name": mat_name,
                "model_spec": model_spec,
                "raw_model_spec": model_spec,
                "unit": unit,
                "unit_price": unit_price,
                "remark": remark,
                "created_by": operator,
                "updated_by": operator,
            })

    # 2. 解析管件与附件单价 Excel (逐行读取，保留原始真实备注)
    if os.path.exists(fitting_excel_path):
        wb_fit = openpyxl.load_workbook(fitting_excel_path, data_only=True)
        ws_fit = wb_fit.active
        for r in range(2, ws_fit.max_row + 1):
            sup_name = _normalize_text(ws_fit.cell(r, 1).value)
            category = _normalize_text(ws_fit.cell(r, 2).value) or "管件"
            mat_name = _normalize_text(ws_fit.cell(r, 3).value)
            raw_spec = _normalize_text(ws_fit.cell(r, 4).value)
            unit = _normalize_text(ws_fit.cell(r, 5).value) or "个"
            raw_price = ws_fit.cell(r, 7).value
            remark = _normalize_text(ws_fit.cell(r, 8).value)

            if not sup_name or not raw_spec:
                continue

            try:
                unit_price = round(float(raw_price or 0), 2)
            except (ValueError, TypeError):
                unit_price = 0.0

            # 构造自解释标准规格型号
            if raw_spec.startswith(mat_name) or (mat_name and mat_name in raw_spec):
                model_spec = raw_spec
            elif mat_name:
                model_spec = f"{mat_name} {raw_spec}".strip()
            else:
                model_spec = f"{category} {raw_spec}".strip()

            entity_id = _resolve_supply_entity_id(sup_name, config)
            all_rows.append({
                "project_key": "insulation_pipe_supply_2026",
                "material_kind": "fitting",
                "supply_entity_id": entity_id,
                "supplier_name": sup_name,
                "category": category,
                "material_name": mat_name or category,
                "model_spec": model_spec,
                "raw_model_spec": raw_spec,
                "unit": unit,
                "unit_price": unit_price,
                "remark": remark,
                "created_by": operator,
                "updated_by": operator,
            })

    # 3. 对同供给方、同规格型号存在重复记录的数据，自动在备注中清晰标注
    from collections import Counter
    group_counts = Counter([(r["supplier_name"], r["model_spec"]) for r in all_rows])
    group_current_index: Dict[tuple, int] = {}

    for row in all_rows:
        key = (row["supplier_name"], row["model_spec"])
        total_cnt = group_counts[key]
        if total_cnt > 1:
            curr_idx = group_current_index.get(key, 0) + 1
            group_current_index[key] = curr_idx
            dup_note = f"同型号多行报价 (第 {curr_idx}/{total_cnt} 笔)"
            orig_rem = row["remark"].strip() if row.get("remark") else ""
            row["remark"] = f"{dup_note}；{orig_rem}" if orig_rem else dup_note

    # 4. 清空现有项目单价并全量逐行插入
    session = SessionLocal()
    try:
        session.execute(text("DELETE FROM tube.tube_material_price WHERE project_key = 'insulation_pipe_supply_2026';"))
        session.execute(text("ALTER SEQUENCE tube.tube_material_price_id_seq RESTART WITH 1;"))

        sql_insert = text("""
            INSERT INTO tube.tube_material_price (
                project_key, material_kind, supply_entity_id, supplier_name,
                category, material_name, model_spec, raw_model_spec,
                unit, unit_price, remark, created_by, created_at,
                updated_by, updated_at
            ) VALUES (
                :project_key, :material_kind, :supply_entity_id, :supplier_name,
                :category, :material_name, :model_spec, :raw_model_spec,
                :unit, :unit_price, :remark, :created_by, NOW(),
                :updated_by, NOW()
            );
        """)

        for row in all_rows:
            session.execute(sql_insert, row)

        session.commit()
    except Exception as e:
        session.rollback()
        raise RuntimeError(f"写入单价数据至数据库失败: {e}") from e
    finally:
        session.close()

    return {
        "success": True,
        "total_inserted": len(all_rows),
        "pipe_file": pipe_excel_path,
        "fitting_file": fitting_excel_path,
    }


def list_material_prices(
    material_kind: Optional[str] = None,
    supplier_name: Optional[str] = None,
    category: Optional[str] = None,
    keyword: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """查询单价字典列表 (默认保温管在前、管件在后)。"""
    ensure_price_table()
    session = SessionLocal()
    try:
        wheres = ["1=1"]
        params: Dict[str, Any] = {}

        if material_kind:
            wheres.append("material_kind = :material_kind")
            params["material_kind"] = material_kind
        if supplier_name:
            wheres.append("supplier_name ILIKE :supplier_name")
            params["supplier_name"] = f"%{supplier_name}%"
        if category:
            wheres.append("category = :category")
            params["category"] = category
        if keyword:
            wheres.append("(model_spec ILIKE :kw OR material_name ILIKE :kw OR raw_model_spec ILIKE :kw OR remark ILIKE :kw)")
            params["kw"] = f"%{keyword}%"

        where_clause = " AND ".join(wheres)
        sql = text(f"""
            SELECT 
                id, project_key, material_kind, supply_entity_id, supplier_name,
                category, material_name, model_spec, raw_model_spec,
                unit, unit_price, remark, created_at, updated_at
            FROM tube.tube_material_price
            WHERE {where_clause}
            ORDER BY material_kind DESC, supplier_name ASC, category ASC, id ASC
        """)
        rows = session.execute(sql, params).mappings().all()
        return [dict(r) for r in rows]
    finally:
        session.close()
