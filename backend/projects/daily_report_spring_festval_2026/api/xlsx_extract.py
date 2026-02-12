# -*- coding: utf-8 -*-
"""
春节简化日报：上传 xlsx 并提取 byDate JSON。
"""

from __future__ import annotations

import io
import math
import re
import ast
from datetime import date, datetime
import json
from typing import Any, Dict, List, Optional, Set, Tuple

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from backend.services.project_data_paths import resolve_project_runtime_path

router = APIRouter(tags=["spring_festival"])
PROJECT_KEY = "daily_report_spring_festval_2026"
LATEST_EXTRACT_FILENAME = "spring_festival_latest_extract.json"

HEADER_LABELS = {"本期", "同期", "差异"}
DATE_TEXT_RE = re.compile(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$")
PERCENT_RE = re.compile(r"^-?\d+(\.\d+)?%$")
NUMBER_RE = re.compile(r"^-?\d+(\.\d+)?$")
CELL_REF_RE = re.compile(r"\b([A-Z]{1,3})(\d+)\b")
ALLOWED_EXPR_RE = re.compile(r"^[\d\.\+\-\*\/\(\)\s]+$")


def _parse_bool(value: Optional[str], default: bool) -> bool:
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _normalize_metric_name(value: Any) -> Optional[str]:
    if value is None:
        return None
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _to_number_maybe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
            return None
        return float(value) if isinstance(value, int) else value
    if isinstance(value, (datetime, date)):
        return value
    if not isinstance(value, str):
        return value
    raw = value.strip()
    if not raw:
        return None
    if raw in {"-", "—", "N/A"}:
        return None
    cleaned = raw.replace(",", "")
    if PERCENT_RE.match(cleaned):
        number = float(cleaned[:-1])
        return number / 100 if math.isfinite(number) else None
    if NUMBER_RE.match(cleaned):
        number = float(cleaned)
        return number if math.isfinite(number) else None
    return value


def _excel_col_to_index(col_letters: str) -> int:
    col = 0
    for char in col_letters:
        col = col * 26 + (ord(char) - ord("A") + 1)
    return col - 1


def _safe_eval_expression(expression: str) -> Optional[float]:
    try:
        tree = ast.parse(expression, mode="eval")
    except Exception:
        return None

    def _eval_node(node) -> Optional[float]:
        if isinstance(node, ast.Expression):
            return _eval_node(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return float(node.value)
            return None
        if isinstance(node, ast.UnaryOp):
            val = _eval_node(node.operand)
            if val is None:
                return None
            if isinstance(node.op, ast.UAdd):
                return val
            if isinstance(node.op, ast.USub):
                return -val
            return None
        if isinstance(node, ast.BinOp):
            left = _eval_node(node.left)
            right = _eval_node(node.right)
            if left is None or right is None:
                return None
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                if right == 0:
                    return None
                return left / right
            return None
        return None

    value = _eval_node(tree)
    if value is None or not math.isfinite(value):
        return None
    return value


def _try_eval_formula(
    matrix: List[List[Any]],
    formula_text: str,
    cache: Dict[Tuple[int, int], Optional[float]],
    visiting: Set[Tuple[int, int]],
) -> Optional[float]:
    expr = formula_text.strip()
    if expr.startswith("="):
        expr = expr[1:].strip()
    if not expr:
        return None

    upper = expr.upper()
    if upper.startswith("IF(") or upper.startswith("IFERROR("):
        return None
    if ":" in expr:
        return None

    def _replace_ref(match: re.Match[str]) -> str:
        col_letters = match.group(1)
        row_number = int(match.group(2))
        row_index = row_number - 1
        col_index = _excel_col_to_index(col_letters)
        value = _resolve_cell_number(matrix, row_index, col_index, cache, visiting)
        if value is None:
            return "0"
        return str(value)

    normalized = CELL_REF_RE.sub(_replace_ref, expr)
    if not ALLOWED_EXPR_RE.match(normalized):
        return None
    return _safe_eval_expression(normalized)


def _resolve_cell_number(
    matrix: List[List[Any]],
    row_index: int,
    col_index: int,
    cache: Dict[Tuple[int, int], Optional[float]],
    visiting: Set[Tuple[int, int]],
) -> Optional[float]:
    key = (row_index, col_index)
    if key in cache:
        return cache[key]
    if key in visiting:
        return None
    if row_index < 0 or col_index < 0:
        return None
    if row_index >= len(matrix):
        return None
    row = matrix[row_index]
    if col_index >= len(row):
        return None

    visiting.add(key)
    raw_value = row[col_index]
    parsed = _to_number_maybe(raw_value)
    result: Optional[float] = None
    if isinstance(parsed, (int, float)):
        result = float(parsed)
    elif isinstance(raw_value, str) and raw_value.strip().startswith("="):
        result = _try_eval_formula(matrix, raw_value, cache, visiting)
    visiting.discard(key)
    cache[key] = result
    return result


def _to_number_or_formula(
    matrix: List[List[Any]],
    raw_value: Any,
    row_index: int,
    col_index: int,
    cache: Dict[Tuple[int, int], Optional[float]],
) -> Any:
    parsed = _to_number_maybe(raw_value)
    if isinstance(parsed, (int, float)):
        return float(parsed)
    if isinstance(raw_value, str) and raw_value.strip().startswith("="):
        value = _resolve_cell_number(matrix, row_index, col_index, cache, set())
        if value is not None:
            return value
    return parsed


def _to_iso_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isinf(value) or math.isnan(value)):
            return None
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date().isoformat()
            if isinstance(converted, date):
                return converted.isoformat()
        except Exception:
            return str(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        matched = DATE_TEXT_RE.match(text)
        if matched:
            year, month, day = matched.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"
        return text
    return str(value)


def _fill_merges(matrix: List[List[Any]], worksheet) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        top_left_value = matrix[min_row - 1][min_col - 1]
        for row_index in range(min_row - 1, max_row):
            for col_index in range(min_col - 1, max_col):
                current = matrix[row_index][col_index]
                if current in {None, ""}:
                    matrix[row_index][col_index] = top_left_value


def _find_header_row(matrix: List[List[Any]]) -> int:
    max_scan = min(len(matrix), 80)
    for row_index in range(max_scan):
        row = matrix[row_index]
        counter = sum(1 for item in row if item in HEADER_LABELS)
        if counter >= 6 and all(label in row for label in HEADER_LABELS):
            return row_index
    return -1


def _set_metric(
    by_date: Dict[str, Dict[str, Dict[str, Any]]],
    date_key: str,
    scope: str,
    metric: str,
    payload: Dict[str, Any],
) -> None:
    date_bucket = by_date.setdefault(date_key, {})
    scope_bucket = date_bucket.setdefault(scope, {})
    existing = scope_bucket.get(metric)
    if existing is None:
        scope_bucket[metric] = payload
        return
    if isinstance(existing, list):
        existing.append(payload)
        return
    scope_bucket[metric] = [existing, payload]


def _extract_bydate_payload(
    worksheet,
    *,
    keep_diff_cell: bool,
    compute_diff: bool,
    normalize_metric: bool,
) -> Dict[str, Any]:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    matrix: List[List[Any]] = [
        [worksheet.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    _fill_merges(matrix, worksheet)

    sub_header_row = _find_header_row(matrix)
    if sub_header_row < 0:
        raise HTTPException(status_code=400, detail="找不到包含“本期/同期/差异”的表头行。")
    date_row = sub_header_row - 1
    if date_row < 0:
        raise HTTPException(status_code=400, detail="表头上一行（日期行）不存在，表结构不符合预期。")

    header = matrix[sub_header_row]
    dates: List[str] = []
    groups: List[Dict[str, Any]] = []
    col_index = 0
    while col_index + 2 < len(header):
        if (
            header[col_index] == "本期"
            and header[col_index + 1] == "同期"
            and header[col_index + 2] == "差异"
        ):
            date_value = matrix[date_row][col_index] if date_row < len(matrix) else None
            date_key = _to_iso_date(date_value) or f"date_col_{col_index + 1}"
            dates.append(date_key)
            groups.append({"date": date_key, "col_start": col_index})
            col_index += 3
            continue
        col_index += 1

    if not groups:
        raise HTTPException(status_code=400, detail="未识别到日期列组（本期/同期/差异）。")

    by_date: Dict[str, Dict[str, Dict[str, Any]]] = {date_key: {} for date_key in dates}
    number_cache: Dict[Tuple[int, int], Optional[float]] = {}

    for row_index in range(sub_header_row + 1, len(matrix)):
        row = matrix[row_index]
        has_any = any(
            cell is not None and str(cell).strip() != ""
            for cell in row
        )
        if not has_any:
            continue

        scope = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        metric_raw = row[1] if len(row) > 1 else None
        unit = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
        if not scope or metric_raw is None or str(metric_raw).strip() == "":
            continue

        metric = (
            _normalize_metric_name(metric_raw)
            if normalize_metric
            else str(metric_raw)
        )
        if not metric:
            continue

        for group in groups:
            start = group["col_start"]
            current_raw = row[start] if start < len(row) else None
            prior_raw = row[start + 1] if start + 1 < len(row) else None
            diff_cell = row[start + 2] if start + 2 < len(row) else None

            current = _to_number_or_formula(
                matrix,
                current_raw,
                row_index,
                start,
                number_cache,
            )
            prior = _to_number_or_formula(
                matrix,
                prior_raw,
                row_index,
                start + 1,
                number_cache,
            )
            diff = None
            if (
                compute_diff
                and isinstance(current, (int, float))
                and isinstance(prior, (int, float))
            ):
                if prior == 0:
                    diff = 0 if current == 0 else None
                else:
                    diff = (current - prior) / prior

            payload: Dict[str, Any] = {
                "unit": unit,
                "current": current if current is not None else None,
                "prior": prior if prior is not None else None,
                "diff": diff,
            }
            if keep_diff_cell:
                payload["diffCell"] = diff_cell

            is_empty = (
                payload["current"] is None
                and payload["prior"] is None
                and payload["diff"] is None
                and (not keep_diff_cell or payload.get("diffCell") is None)
            )
            if is_empty:
                continue

            _set_metric(by_date, group["date"], scope, metric, payload)

    return {
        "meta": {
            "sheetName": worksheet.title,
            "generatedAt": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
            "header": {
                "dateRow": date_row,
                "subHeaderRow": sub_header_row,
            },
        },
        "dates": dates,
        "byDate": by_date,
    }


@router.post("/spring-festival/extract-json", summary="上传 xlsx 并提取 byDate JSON")
async def upload_and_extract_json(
    file: UploadFile = File(..., description="春节简化日报 xlsx 文件"),
    sheet_name: Optional[str] = Form(default=None),
    keep_diff_cell: Optional[str] = Form(default="true"),
    compute_diff: Optional[str] = Form(default="true"),
    normalize_metric: Optional[str] = Form(default="true"),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm/.xltx/.xltm 文件。")

    binary = await file.read()
    if not binary:
        raise HTTPException(status_code=400, detail="上传文件为空。")

    try:
        workbook = load_workbook(
            filename=io.BytesIO(binary),
            data_only=False,
            read_only=False,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 Excel 失败：{exc}") from exc

    worksheet = None
    if sheet_name:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if worksheet is None:
            raise HTTPException(
                status_code=400,
                detail=f"工作表不存在：{sheet_name}，可用工作表：{', '.join(workbook.sheetnames)}",
            )
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    payload = _extract_bydate_payload(
        worksheet,
        keep_diff_cell=_parse_bool(keep_diff_cell, True),
        compute_diff=_parse_bool(compute_diff, True),
        normalize_metric=_parse_bool(normalize_metric, True),
    )
    payload["meta"]["fileName"] = filename
    payload["meta"]["sheetCandidates"] = list(workbook.sheetnames)
    runtime_file = resolve_project_runtime_path(PROJECT_KEY, LATEST_EXTRACT_FILENAME)
    runtime_file.parent.mkdir(parents=True, exist_ok=True)
    runtime_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


@router.get("/spring-festival/latest-json", summary="读取最近一次提取 JSON")
def get_latest_extract_json():
    runtime_file = resolve_project_runtime_path(PROJECT_KEY, LATEST_EXTRACT_FILENAME)
    if not runtime_file.exists():
        return {"ok": False, "message": "暂无最近提取数据"}
    try:
        payload = json.loads(runtime_file.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取最近提取数据失败：{exc}") from exc
    return {"ok": True, "payload": payload}
